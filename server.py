"""FastAPI backend for the OutboundAI dashboard."""

import asyncio
import base64
import csv
import hashlib
import hmac
from datetime import datetime
from io import BytesIO, StringIO
import json
import logging
import os
import random
import ssl
import time
import uuid
import certifi
import aiohttp
from pathlib import Path
from urllib.parse import urlparse, unquote
from typing import Optional
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel

_orig_ssl = ssl.create_default_context

def _certifi_ssl(purpose=ssl.Purpose.SERVER_AUTH, **kwargs):
    if not kwargs.get("cafile") and not kwargs.get("capath") and not kwargs.get("cadata"):
        kwargs["cafile"] = certifi.where()
    return _orig_ssl(purpose, **kwargs)

ssl.create_default_context = _certifi_ssl

from db import (
    ConfigError, CRM_TERMINAL_STATUSES, DuplicateContactError,
    cancel_appointment, clear_all_test_data, clear_appointments,
    clear_call_logs, clear_campaigns, clear_contact_memory, clear_error_logs,
    clear_errors, create_agent_profile, create_campaign, delete_agent_profile,
    delete_campaign, get_agent_profile,
    get_all_agent_profiles, get_all_appointments, get_all_calls,
    get_all_campaigns, get_all_settings, get_calls_by_phone, get_campaign,
    get_call_logs_for_export, get_contacts, get_crm_contact_detail, get_crm_contacts,
    get_crm_summary, get_lead_statuses, get_crm_contact_by_phone,
    get_logs, get_recording_storage_stats, get_recordings_for_cleanup,
    get_setting, get_stats, init_db, log_error, mark_recording_deleted,
    normalize_phone, _tz_today, upsert_crm_lead,
    save_settings, set_default_agent_profile, set_setting,
    add_lead_status, delete_lead_status, update_agent_profile, update_call_notes,
    update_campaign_contacts, update_campaign_run_stats, update_campaign_status,
    update_crm_contact_followup, update_crm_contact_full, update_crm_contact_notes,
    update_crm_contact_status,
    get_knowledge_base, save_knowledge_base, get_kb_section, save_kb_section,
    _KB_SECTIONS,
)
from prompts import (
    DEFAULT_SYSTEM_PROMPT,
    PROMPT_TYPES,
    get_default_prompt,
    get_prompt_label,
    build_prompt_for_type,
    build_knowledge_context,
    get_kb_prompt_prefix,
)
from whatsapp import (
    WA_SETTINGS_KEYS, AUTOMATION_EVENT_TYPES, AUTOMATION_ACTION_TYPES,
    get_wa_settings_masked, save_wa_settings, get_wa_health,
    send_whatsapp_template, get_whatsapp_logs,
    get_automation_rules, save_automation_rules, find_automation_rule, source_to_event_type,
    execute_automation_rule,
    insert_automation_action, get_automation_actions, update_automation_action_status,
    run_due_automation_actions,
    send_callback_confirmation, send_appointment_confirmation, send_showroom_visit_confirmation,
    handle_call_outcome_whatsapp_fallback,
    # Phase 8 — WhatsApp Inbox
    get_conversations, get_conversation_by_id, get_messages, patch_conversation,
    save_wa_message, send_whatsapp_text, is_whatsapp_service_window_open,
    get_or_create_conversation, update_conversation_last_message,
    parse_webhook_messages, handle_inbound_whatsapp_message,
)

load_dotenv(".env", override=False)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("server")

init_db()

try:
    from apscheduler.schedulers.asyncio import AsyncIOScheduler
    from apscheduler.triggers.cron import CronTrigger
    _scheduler = AsyncIOScheduler()
except ImportError:
    _scheduler = None
    logger.warning("APScheduler not installed — campaign scheduling disabled")

app = FastAPI(title="OutboundAI Dashboard", version="1.0.0")

SESSION_COOKIE_NAME = "aicalling_admin_session"
SESSION_TTL_SECONDS = 12 * 60 * 60
PUBLIC_API_ROUTES = {
    ("GET", "/api/health"),
    ("POST", "/api/auth/login"),
    ("GET", "/api/auth/me"),
    ("POST", "/api/auth/logout"),
    ("GET", "/api/whatsapp/webhook"),   # Phase 8 — verify token
    ("POST", "/api/whatsapp/webhook"),  # Phase 8 — incoming messages
    ("POST", "/api/leads/incoming"),    # n8n / Facebook / website lead intake
}


def _admin_username() -> str:
    return os.getenv("ADMIN_USERNAME", "")


def _admin_password() -> str:
    return os.getenv("ADMIN_PASSWORD", "")


def _session_secret() -> str:
    return os.getenv("SESSION_SECRET", "")


def _admin_config_error() -> Optional[str]:
    if not (_admin_username() and _admin_password()):
        return "Admin login is not configured. Set ADMIN_USERNAME and ADMIN_PASSWORD."
    if not _session_secret():
        return "Admin login is not configured. Set SESSION_SECRET."
    return None


def _b64encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def _b64decode(data: str) -> bytes:
    padded = data + "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode(padded.encode("ascii"))


def _sign_session(payload: str) -> str:
    return hmac.new(_session_secret().encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()


def _make_session_token(username: str) -> str:
    payload = _b64encode(json.dumps({"u": username, "exp": int(time.time()) + SESSION_TTL_SECONDS}, separators=(",", ":")).encode("utf-8"))
    return f"{payload}.{_sign_session(payload)}"


def _read_session(request: Request) -> Optional[str]:
    token = request.cookies.get(SESSION_COOKIE_NAME, "")
    if not token or "." not in token or not _session_secret():
        return None
    payload, signature = token.rsplit(".", 1)
    if not hmac.compare_digest(signature, _sign_session(payload)):
        return None
    try:
        data = json.loads(_b64decode(payload).decode("utf-8"))
    except Exception:
        return None
    if int(data.get("exp", 0)) < int(time.time()):
        return None
    username = data.get("u")
    return username if username and hmac.compare_digest(username, _admin_username()) else None


def _request_is_https(request: Request) -> bool:
    proto = request.headers.get("x-forwarded-proto", "").split(",")[0].strip().lower()
    return request.url.scheme == "https" or proto == "https"


@app.middleware("http")
async def _admin_auth_middleware(request: Request, call_next):
    if request.url.path.startswith("/api/") and (request.method, request.url.path) not in PUBLIC_API_ROUTES:
        if not _read_session(request):
            return JSONResponse(status_code=401, content={"error": "Unauthorized"})
    return await call_next(request)


@app.exception_handler(ConfigError)
async def _config_error_handler(request: Request, exc: ConfigError):
    return JSONResponse(status_code=503, content={"error": str(exc)})


@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: Request, exc: Exception):
    msg = str(exc)
    # Map the most common Supabase misconfig to a clean 503 instead of a 500 stack trace.
    if "supabase_key is required" in msg or "supabase_url is required" in msg.lower():
        return JSONResponse(
            status_code=503,
            content={"error": "Supabase not configured. Set SUPABASE_URL and SUPABASE_SERVICE_KEY env vars."},
        )
    logger.exception("Unhandled error on %s: %s", request.url.path, exc)
    return JSONResponse(status_code=500, content={"error": msg})


@app.on_event("startup")
async def _startup():
    if _scheduler:
        _scheduler.start()
        await _reschedule_all_campaigns()
        await _schedule_recording_cleanup()
        # Automation action runner — every 60 seconds
        try:
            from apscheduler.triggers.interval import IntervalTrigger
            _scheduler.add_job(
                lambda: asyncio.create_task(run_due_automation_actions()),
                trigger=IntervalTrigger(seconds=60),
                id="automation_runner",
                replace_existing=True,
            )
        except Exception as _e:
            logger.warning("Automation runner schedule failed: %s", _e)


@app.on_event("shutdown")
async def _shutdown():
    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)


async def eff(key: str) -> str:
    val = await get_setting(key, "")
    return val if val else os.getenv(key, "")


class CallRequest(BaseModel):
    phone: str
    lead_name: str = "there"
    business_name: str = "our company"
    service_type: str = "our service"
    system_prompt: Optional[str] = None
    agent_profile_id: Optional[str] = None
    call_type: Optional[str] = None


class AgentProfileRequest(BaseModel):
    name: str
    voice: str = "Aoede"
    model: str = "gemini-3.1-flash-live-preview"
    system_prompt: Optional[str] = None
    enabled_tools: str = "[]"
    is_default: bool = False


class PromptRequest(BaseModel):
    prompt: str


class AiPromptRequest(BaseModel):
    prompt: str
    is_default: bool = False


class SettingsRequest(BaseModel):
    settings: dict


class AuthRequest(BaseModel):
    username: str
    password: str


class NotesRequest(BaseModel):
    notes: str


class CampaignRequest(BaseModel):
    name: str
    contacts: list
    schedule_type: str = "once"
    schedule_time: str = "09:00"
    call_delay_seconds: int = 3
    system_prompt: Optional[str] = None
    agent_profile_id: Optional[str] = None


class StatusRequest(BaseModel):
    status: str


class ClearRecordsRequest(BaseModel):
    confirm: Optional[str] = None


class LeadStatusRequest(BaseModel):
    name: str
    color: Optional[str] = None


class CrmStatusRequest(BaseModel):
    crm_status: str
    custom_status: Optional[str] = None


class CrmFollowupRequest(BaseModel):
    next_followup_at: Optional[str] = None


class CrmNotesRequest(BaseModel):
    crm_notes: str = ""


class CrmLeadRequest(BaseModel):
    phone_number: str
    lead_name: str
    email: Optional[str] = None
    city: Optional[str] = None
    location: Optional[str] = None
    requirement: Optional[str] = None
    budget: Optional[str] = None
    source: Optional[str] = "manual"
    business_name: Optional[str] = None
    campaign_name: Optional[str] = None
    service_type: Optional[str] = None
    crm_status: Optional[str] = None
    custom_status: Optional[str] = None
    crm_notes: Optional[str] = None
    next_followup_at: Optional[str] = None
    assigned_to: Optional[str] = None


class CrmBulkImportRequest(BaseModel):
    leads: list
    import_source: Optional[str] = "api"


class CrmLeadUpdateRequest(BaseModel):
    lead_name: Optional[str] = None
    email: Optional[str] = None
    city: Optional[str] = None
    location: Optional[str] = None
    requirement: Optional[str] = None
    budget: Optional[str] = None
    source: Optional[str] = None
    business_name: Optional[str] = None
    campaign_name: Optional[str] = None
    service_type: Optional[str] = None
    crm_status: Optional[str] = None
    custom_status: Optional[str] = None
    crm_notes: Optional[str] = None
    next_followup_at: Optional[str] = None
    assigned_to: Optional[str] = None


class CrmCallSelectedRequest(BaseModel):
    phones: list
    business_name: str = "our company"
    service_type: str = "our service"
    campaign_name: str = "Selected CRM Leads"
    call_delay_seconds: int = 15
    agent_profile_id: Optional[str] = None
    system_prompt: Optional[str] = None


class RecordingCleanupRequest(BaseModel):
    confirm: Optional[str] = None


class BatchCallRequest(BaseModel):
    contacts: list  # [{phone, lead_name, business_name, service_type}]
    call_delay_seconds: int = 5
    agent_profile_id: Optional[str] = None
    system_prompt: Optional[str] = None
    batch_name: Optional[str] = None


class OutboundScheduleRequest(BaseModel):
    OUTBOUND_TIMEZONE: Optional[str] = None
    OUTBOUND_START_TIME: Optional[str] = None
    OUTBOUND_END_TIME: Optional[str] = None
    OUTBOUND_ALLOWED_DAYS: Optional[str] = None
    OUTBOUND_CALLING_ENABLED: Optional[str] = None


class WaSettingsRequest(BaseModel):
    settings: dict


class WaSendTemplateRequest(BaseModel):
    phone: str
    template_name: str
    language: str = "en"
    parameters: Optional[list] = None


class AutomationRulesRequest(BaseModel):
    rules: list


class LeadIncomingRequest(BaseModel):
    name: Optional[str] = None
    phone: str
    email: Optional[str] = None
    source: Optional[str] = "api"
    city: Optional[str] = None
    service: Optional[str] = None
    message: Optional[str] = None
    business_name: Optional[str] = None
    campaign_name: Optional[str] = None
    extra: Optional[dict] = None


class AutomationTestRequest(BaseModel):
    phone: str
    event_type: str
    source: Optional[str] = None
    lead_name: Optional[str] = "Test Lead"
    dry_run: bool = True


class WaConfirmRequest(BaseModel):
    phone: str
    lead_name: Optional[str] = None
    business_name: Optional[str] = None
    service_type: Optional[str] = None
    appointment_date: Optional[str] = None
    appointment_time: Optional[str] = None
    address: Optional[str] = None
    contact_number: Optional[str] = None


@app.post("/api/auth/login")
async def api_auth_login(req: AuthRequest, request: Request):
    config_error = _admin_config_error()
    if config_error:
        return JSONResponse(status_code=503, content={"error": config_error})
    if not (hmac.compare_digest(req.username, _admin_username()) and hmac.compare_digest(req.password, _admin_password())):
        return JSONResponse(status_code=401, content={"error": "Invalid username or password"})
    response = JSONResponse(content={"success": True})
    response.set_cookie(
        SESSION_COOKIE_NAME,
        _make_session_token(req.username),
        max_age=SESSION_TTL_SECONDS,
        httponly=True,
        secure=_request_is_https(request),
        samesite="lax",
        path="/",
    )
    return response


@app.get("/api/auth/me")
async def api_auth_me(request: Request):
    config_error = _admin_config_error()
    username = _read_session(request)
    payload = {"authenticated": bool(username), "username": username if username else None}
    if config_error:
        payload["configured"] = False
        payload["error"] = config_error
    else:
        payload["configured"] = True
    return payload


@app.post("/api/auth/logout")
async def api_auth_logout(request: Request):
    response = JSONResponse(content={"success": True})
    response.delete_cookie(SESSION_COOKIE_NAME, path="/", samesite="lax", secure=_request_is_https(request))
    return response


@app.get("/api/health")
async def api_health():
    async def status_value(key: str, *fallback_envs: str) -> str:
        for env_key in (key, *fallback_envs):
            val = os.getenv(env_key, "")
            if val:
                return val
        return ""

    async def supabase_status(url: str, key: str) -> tuple[bool, Optional[str]]:
        if not (url and key):
            return False, None
        try:
            timeout = aiohttp.ClientTimeout(total=3)
            headers = {"apikey": key, "Authorization": f"Bearer {key}"}
            async with aiohttp.ClientSession(timeout=timeout) as session:
                async with session.get(f"{url.rstrip('/')}/rest/v1/", headers=headers) as resp:
                    if resp.status in (401, 403):
                        return False, "Invalid API key"
                    return True, None
        except Exception:
            return True, None

    livekit_url = await status_value("LIVEKIT_URL")
    livekit_key = await status_value("LIVEKIT_API_KEY")
    livekit_secret = await status_value("LIVEKIT_API_SECRET")
    google_key = await status_value("GOOGLE_API_KEY")
    supabase_url = await status_value("SUPABASE_URL")
    supabase_key = await status_value("SUPABASE_SERVICE_KEY")
    trunk_id = await status_value("OUTBOUND_TRUNK_ID")
    gemini_model = await status_value("GEMINI_MODEL")
    gemini_voice = await status_value("GEMINI_TTS_VOICE")
    prompt_saved = await status_value("SYSTEM_PROMPT", "system_prompt")
    s3_key = await status_value("S3_ACCESS_KEY_ID", "AWS_ACCESS_KEY_ID")
    s3_secret = await status_value("S3_SECRET_ACCESS_KEY", "AWS_SECRET_ACCESS_KEY")
    s3_bucket = await status_value("S3_BUCKET", "AWS_BUCKET_NAME")
    recording_auto_delete = (await status_value("RECORDING_AUTO_DELETE_ENABLED") or "false").strip().lower() in ("1", "true", "yes", "on")
    try:
        recording_retention_days = max(int(await status_value("RECORDING_RETENTION_DAYS") or "7"), 1)
    except ValueError:
        recording_retention_days = 7
    supabase_configured, supabase_error = await supabase_status(supabase_url, supabase_key)
    response = {
        "status": "ok",
        "livekit_configured": bool(livekit_url and livekit_key and livekit_secret),
        "gemini_configured": bool(google_key),
        "supabase_configured": supabase_configured,
        "trunk_configured": bool(trunk_id),
        "gemini_model_configured": bool(gemini_model),
        "gemini_tts_voice_configured": bool(gemini_voice),
        "prompt_configured": bool(prompt_saved),
        "prompt_mode": "custom" if prompt_saved else "unknown",
        "s3_configured": bool(s3_key and s3_secret and s3_bucket),
        "recording_auto_delete_enabled": recording_auto_delete,
        "recording_retention_days": recording_retention_days,
    }
    if supabase_error:
        response["supabase_error"] = supabase_error
    return response


@app.get("/", response_class=HTMLResponse)
async def serve_dashboard():
    html_path = Path(__file__).parent / "ui" / "index.html"
    if html_path.exists():
        return HTMLResponse(content=html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>Dashboard not found — place index.html in ui/</h1>", status_code=404)


@app.post("/api/call")
async def api_dispatch_call(req: CallRequest):
    # ── Outbound calling-window guard ──────────────────────────────────────
    if not await _is_outbound_allowed():
        raise HTTPException(403, await _outbound_window_error())
    url = await eff("LIVEKIT_URL")
    key = await eff("LIVEKIT_API_KEY")
    secret = await eff("LIVEKIT_API_SECRET")
    if not all([url, key, secret]):
        raise HTTPException(400, "LiveKit credentials not configured. Go to Settings → LiveKit.")
    # Pre-flight: without these the agent will silently fail mid-call.
    if not await eff("OUTBOUND_TRUNK_ID"):
        raise HTTPException(
            400,
            "OUTBOUND_TRUNK_ID is not set. Either click 'Create SIP Trunk' in Settings → Vobiz, "
            "or paste an existing trunk ID into your Coolify env vars and redeploy.",
        )
    if not await eff("GOOGLE_API_KEY"):
        raise HTTPException(400, "GOOGLE_API_KEY is not set — the Gemini Live agent cannot connect.")
    phone = req.phone.strip()
    if not phone.startswith("+"):
        raise HTTPException(400, "Phone must be in E.164 format: +919876543210")

    effective_prompt = req.system_prompt
    effective_voice = effective_model = effective_tools = None
    if req.agent_profile_id:
        profile = await get_agent_profile(req.agent_profile_id)
        if profile:
            if not effective_prompt and profile.get("system_prompt"):
                effective_prompt = profile["system_prompt"]
            effective_voice = profile.get("voice")
            effective_model = profile.get("model")
            effective_tools = profile.get("enabled_tools")
    if not effective_prompt:
        # Try legacy global prompt first, then resolve by call_type
        effective_prompt = await get_setting("system_prompt", "") or None
    if not effective_prompt:
        call_type = (req.call_type or "welcome_call").strip()
        effective_prompt = await resolve_ai_prompt(
            call_type=call_type,
            lead_name=req.lead_name,
            business_name=req.business_name,
            service_type=req.service_type,
        )

    room_name = f"call-{phone.replace('+', '')}-{random.randint(1000, 9999)}"
    metadata = {"phone_number": phone, "lead_name": req.lead_name, "business_name": req.business_name, "service_type": req.service_type, "system_prompt": effective_prompt}
    if effective_voice:
        metadata["voice_override"] = effective_voice
    if effective_model:
        metadata["model_override"] = effective_model
    if effective_tools:
        metadata["tools_override"] = effective_tools
    try:
        from livekit import api as lk_api
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        session = aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=ctx))
        lk = lk_api.LiveKitAPI(url=url, api_key=key, api_secret=secret, session=session)
        await lk.room.create_room(lk_api.CreateRoomRequest(name=room_name, empty_timeout=300, max_participants=5))
        await lk.agent_dispatch.create_dispatch(lk_api.CreateAgentDispatchRequest(agent_name="outbound-caller", room=room_name, metadata=json.dumps(metadata)))
        await lk.aclose()
        await session.close()
        await log_error("server", f"Call dispatched to {phone}", f"room={room_name}", "info")
        return {"status": "dispatched", "room": room_name, "phone": phone}
    except Exception as exc:
        logger.error("Dispatch error: %s", exc)
        raise HTTPException(500, f"Dispatch failed: {exc}")


@app.get("/api/calls")
async def api_get_calls(page: int = 1, limit: int = 20):
    return await get_all_calls(page=page, limit=limit)


@app.patch("/api/calls/{call_id}/notes")
async def api_update_notes(call_id: str, req: NotesRequest):
    ok = await update_call_notes(call_id, req.notes)
    if not ok:
        raise HTTPException(404, "Call not found")
    return {"status": "updated"}


EXPORT_COLUMNS = [
    ("Call Date/Time", "timestamp"),
    ("Lead Name", "lead_name"),
    ("Phone Number", "phone_number"),
    ("Business Name", "business_name"),
    ("Service Type", "service_type"),
    ("Outcome", "outcome"),
    ("Reason", "reason"),
    ("Duration Seconds", "duration_seconds"),
    ("Notes", "notes"),
    ("Recording URL", "recording_url"),
]


def _recording_link_value(row: dict) -> str:
    """Plain value (URL) for CSV cells and unlinked XLSX fallback."""
    if row.get("recording_deleted"):
        return "Recording Deleted"
    url = row.get("recording_url")
    return url if url else "No Recording"


def _export_filters(date_from: Optional[str], date_to: Optional[str], outcome: Optional[str], phone: Optional[str]) -> dict:
    return {
        "date_from": date_from or "",
        "date_to": date_to or "",
        "outcome": outcome or "",
        "phone": phone or "",
    }


def _export_value(row: dict, key: str) -> str:
    value = row.get(key)
    return "" if value is None else str(value)


def _export_cell_value(row: dict, label: str, key: str) -> str:
    if label == "Recording URL":
        if row.get("recording_deleted"):
            return ""
        return row.get("recording_url") or ""
    return _export_value(row, key)


@app.get("/api/export/calls.csv")
async def api_export_calls_csv(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outcome: Optional[str] = None,
    phone: Optional[str] = None,
):
    try:
        rows = await get_call_logs_for_export(_export_filters(date_from, date_to, outcome, phone))
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow([label for label, _ in EXPORT_COLUMNS])
        for row in rows:
            writer.writerow([
                _export_cell_value(row, label, key)
                for label, key in EXPORT_COLUMNS
            ])
        out.seek(0)
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="call_logs_export.csv"'},
        )
    except Exception as exc:
        await log_error("server", "CSV export failed", str(exc), "error")
        return JSONResponse(status_code=500, content={"error": f"CSV export failed: {exc}"})


@app.get("/api/export/calls.xlsx")
async def api_export_calls_xlsx(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outcome: Optional[str] = None,
    phone: Optional[str] = None,
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        rows = await get_call_logs_for_export(_export_filters(date_from, date_to, outcome, phone))
        wb = Workbook()
        ws = wb.active
        ws.title = "Call Logs"
        headers = [label for label, _ in EXPORT_COLUMNS]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in rows:
            ws.append([
                _export_cell_value(row, label, key)
                for label, key in EXPORT_COLUMNS
            ])

        for column in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 14), 60)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="call_logs_export.xlsx"'},
        )
    except Exception as exc:
        await log_error("server", "Excel export failed", str(exc), "error")
        return JSONResponse(status_code=500, content={"error": f"Excel export failed: {exc}"})


@app.get("/api/stats")
async def api_get_stats():
    return await get_stats()


async def _recording_retention_days() -> int:
    raw = await get_setting("RECORDING_RETENTION_DAYS", "7")
    try:
        return max(int(raw), 1)
    except (TypeError, ValueError):
        return 7


async def _recording_auto_delete_enabled() -> bool:
    raw = (await get_setting("RECORDING_AUTO_DELETE_ENABLED", "false")).strip().lower()
    return raw in ("1", "true", "yes", "on")


async def _recording_cleanup_time() -> str:
    return await get_setting("RECORDING_CLEANUP_TIME", "02:00") or "02:00"


def _derive_recording_key(recording_url: str, bucket: str) -> Optional[str]:
    if not recording_url:
        return None
    parsed = urlparse(recording_url)
    if parsed.scheme == "s3":
        key = parsed.path.lstrip("/")
    else:
        path = unquote(parsed.path or "").lstrip("/")
        prefix = f"{bucket}/"
        key = path[len(prefix):] if bucket and path.startswith(prefix) else path
    if not key or key.startswith("/") or ".." in key.split("/"):
        return None
    return key


async def _delete_recording_object(object_key: str) -> None:
    import boto3

    aws_key = await eff("S3_ACCESS_KEY_ID") or os.getenv("AWS_ACCESS_KEY_ID", "")
    aws_secret = await eff("S3_SECRET_ACCESS_KEY") or os.getenv("AWS_SECRET_ACCESS_KEY", "")
    bucket = await eff("S3_BUCKET") or os.getenv("AWS_BUCKET_NAME", "")
    endpoint = await eff("S3_ENDPOINT_URL") or os.getenv("S3_ENDPOINT", "")
    region = await eff("S3_REGION") or os.getenv("AWS_REGION", "ap-northeast-1")
    if not (aws_key and aws_secret and bucket):
        raise RuntimeError("S3/R2 credentials are not configured")
    client = boto3.client(
        "s3",
        aws_access_key_id=aws_key,
        aws_secret_access_key=aws_secret,
        endpoint_url=endpoint or None,
        region_name=region,
    )
    client.delete_object(Bucket=bucket, Key=object_key)


async def _cleanup_old_recordings() -> dict:
    retention_days = await _recording_retention_days()
    rows = await get_recordings_for_cleanup(retention_days)
    bucket = await eff("S3_BUCKET") or os.getenv("AWS_BUCKET_NAME", "")
    deleted = failed = 0
    for row in rows:
        object_key = row.get("recording_object_key") or _derive_recording_key(row.get("recording_url", ""), bucket)
        if not object_key:
            failed += 1
            continue
        try:
            await _delete_recording_object(object_key)
            if await mark_recording_deleted(row["id"]):
                deleted += 1
            else:
                failed += 1
        except Exception as exc:
            failed += 1
            await log_error("server", "Recording delete failed", f"id={row.get('id')} key={object_key}: {exc}", "error")
    await log_error("server", "Recording cleanup summary", f"deleted={deleted}, failed={failed}, retention_days={retention_days}", "info")
    return {"deleted": deleted, "failed": failed}


@app.get("/api/recordings/storage")
async def api_recordings_storage():
    stats = await get_recording_storage_stats()
    retention_days = await _recording_retention_days()
    auto_delete_enabled = await _recording_auto_delete_enabled()
    return {**stats, "retention_days": retention_days, "auto_delete_enabled": auto_delete_enabled}


@app.post("/api/recordings/cleanup")
async def api_recordings_cleanup(req: Optional[RecordingCleanupRequest] = None):
    if not req or req.confirm != "DELETE_OLD_RECORDINGS":
        return JSONResponse(status_code=400, content={"success": False, "error": "Confirmation required"})
    result = await _cleanup_old_recordings()
    return {"success": True, **result, "message": "Old recordings cleaned up"}


@app.get("/api/appointments")
async def api_get_appointments(date: Optional[str] = None):
    return await get_all_appointments(date_filter=date)


@app.delete("/api/appointments/{appointment_id}")
async def api_cancel_appointment(appointment_id: str):
    ok = await cancel_appointment(appointment_id)
    if not ok:
        raise HTTPException(404, "Appointment not found or already cancelled")
    return {"status": "cancelled"}


@app.get("/api/prompt")
async def api_get_prompt():
    saved = await get_setting("system_prompt", "")
    return {"prompt": saved or DEFAULT_SYSTEM_PROMPT, "is_custom": bool(saved)}


@app.post("/api/prompt")
async def api_save_prompt(req: PromptRequest):
    await set_setting("system_prompt", req.prompt)
    return {"status": "saved"}


@app.delete("/api/prompt")
async def api_reset_prompt():
    await set_setting("system_prompt", "")
    return {"status": "reset", "prompt": DEFAULT_SYSTEM_PROMPT}


# ── AI Prompt Manager helpers ────────────────────────────────────────────────

_AI_PROMPT_KEY = "AI_PROMPT_{}"
_AI_PROMPT_DEFAULT_TYPE_KEY = "AI_PROMPT_DEFAULT_TYPE"
_VALID_PROMPT_TYPES = {pt for pt, _, _ in PROMPT_TYPES}


async def resolve_ai_prompt(
    call_type: str = "welcome_call",
    fallback: str = "welcome_call",
    lead_name: str = "there",
    business_name: str = "our company",
    service_type: str = "our service",
) -> str:
    """
    Resolve final prompt for a call.
    Priority: saved AI_PROMPT_{type} → built-in default for type → fallback type.
    KB context is prepended when knowledge base has content.
    Returns an interpolated prompt string ready to pass to Gemini.
    """
    resolved_type = call_type if call_type in _VALID_PROMPT_TYPES else fallback
    saved = await get_setting(_AI_PROMPT_KEY.format(resolved_type), "")
    kb = await get_knowledge_base()
    return build_prompt_for_type(
        prompt_type=resolved_type,
        lead_name=lead_name,
        business_name=business_name,
        service_type=service_type,
        saved_text=saved or None,
        kb=kb,
    )


def _crm_status_to_call_type(crm_status: str) -> str:
    """Map a CRM lead status to the appropriate call type prompt."""
    status = (crm_status or "").strip().lower()
    if "callback" in status:
        return "callback_call"
    if "follow" in status:
        return "followup_call"
    if "appointment" in status:
        return "appointment_confirmation"
    if "re-enquiry" in status or "re_enquiry" in status or "re enquiry" in status:
        return "re_enquiry"
    if "payment" in status:
        return "payment_followup"
    if "missed" in status or "no answer" in status:
        return "missed_call_retry"
    return "welcome_call"


# ── AI Prompt Manager API endpoints ─────────────────────────────────────────

@app.get("/api/ai-prompts")
async def api_list_ai_prompts():
    default_type = await get_setting(_AI_PROMPT_DEFAULT_TYPE_KEY, "welcome_call")
    result = []
    for pt, label, _ in PROMPT_TYPES:
        saved = await get_setting(_AI_PROMPT_KEY.format(pt), "")
        result.append({
            "type": pt,
            "label": label,
            "prompt": saved if saved else get_default_prompt(pt),
            "is_saved": bool(saved),
            "is_default": (pt == default_type),
        })
    return result


@app.get("/api/ai-prompts/resolve")
async def api_resolve_ai_prompt(call_type: str = "welcome_call"):
    if call_type not in _VALID_PROMPT_TYPES:
        raise HTTPException(400, f"Unknown call type: {call_type}. Valid: {sorted(_VALID_PROMPT_TYPES)}")
    saved = await get_setting(_AI_PROMPT_KEY.format(call_type), "")
    prompt = saved if saved else get_default_prompt(call_type)
    return {
        "type": call_type,
        "label": get_prompt_label(call_type),
        "prompt": prompt,
        "is_saved": bool(saved),
    }


@app.get("/api/ai-prompts/{prompt_type}")
async def api_get_ai_prompt(prompt_type: str):
    if prompt_type not in _VALID_PROMPT_TYPES:
        raise HTTPException(404, f"Unknown prompt type: {prompt_type}")
    default_type = await get_setting(_AI_PROMPT_DEFAULT_TYPE_KEY, "welcome_call")
    saved = await get_setting(_AI_PROMPT_KEY.format(prompt_type), "")
    return {
        "type": prompt_type,
        "label": get_prompt_label(prompt_type),
        "prompt": saved if saved else get_default_prompt(prompt_type),
        "is_saved": bool(saved),
        "is_default": (prompt_type == default_type),
        "default_prompt": get_default_prompt(prompt_type),
    }


@app.post("/api/ai-prompts/{prompt_type}")
async def api_save_ai_prompt(prompt_type: str, req: AiPromptRequest):
    if prompt_type not in _VALID_PROMPT_TYPES:
        raise HTTPException(400, f"Unknown prompt type: {prompt_type}")
    await set_setting(_AI_PROMPT_KEY.format(prompt_type), req.prompt)
    if req.is_default:
        await set_setting(_AI_PROMPT_DEFAULT_TYPE_KEY, prompt_type)
    return {"status": "saved", "type": prompt_type, "is_default": req.is_default}


@app.post("/api/ai-prompts/{prompt_type}/reset")
async def api_reset_ai_prompt(prompt_type: str):
    if prompt_type not in _VALID_PROMPT_TYPES:
        raise HTTPException(400, f"Unknown prompt type: {prompt_type}")
    await set_setting(_AI_PROMPT_KEY.format(prompt_type), "")
    return {
        "status": "reset",
        "type": prompt_type,
        "prompt": get_default_prompt(prompt_type),
    }


# ── Knowledge Base API endpoints ─────────────────────────────────────────────

_KB_SAMPLE = {
    "company_profile": {
        "business_name": "Demo Business",
        "legal_name": "",
        "short_description": "Professional customer support and appointment services.",
        "about_us": "We provide professional customer support and appointment booking services for individuals and businesses.",
        "industry_type": "Local Services",
        "owner_name": "",
        "website": "",
        "email": "info@demobusiness.com",
        "phone": "",
        "whatsapp_number": "",
    },
    "working_hours": {
        "timezone": "Asia/Kolkata",
        "opening_days": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
        "opening_time": "09:00",
        "closing_time": "18:00",
        "holiday_notes": "Closed on national holidays.",
        "emergency_support_available": False,
    },
    "faqs": [
        {"question": "What are your working hours?", "answer": "We are open Monday to Saturday, 9 AM to 6 PM.", "category": "General", "language": "English", "active": True},
        {"question": "How can I book an appointment?", "answer": "You can book an appointment by calling us or through our AI assistant.", "category": "Booking", "language": "English", "active": True},
        {"question": "What is your cancellation policy?", "answer": "You can cancel or reschedule up to 24 hours before your appointment at no charge.", "category": "Policy", "language": "English", "active": True},
    ],
    "appointment_rules": {
        "appointment_required": True,
        "allow_same_day_booking": True,
        "appointment_duration_minutes": 30,
        "appointment_buffer_minutes": 15,
        "default_visit_type": "phone_consultation",
        "confirmation_required": True,
        "reminder_before_hours": 24,
    },
}


@app.get("/api/knowledge-base")
async def api_get_knowledge_base():
    return await get_knowledge_base()


@app.post("/api/knowledge-base")
async def api_save_knowledge_base(req: Request):
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    if not isinstance(data, dict):
        raise HTTPException(400, "Body must be a JSON object")
    saved = await save_knowledge_base(data)
    return {"status": "saved", "sections_saved": [k for k in data if k in _KB_SECTIONS]}


@app.get("/api/knowledge-base/search")
async def api_search_knowledge_base(q: str = Query(default="")):
    if not q or not q.strip():
        return {"results": []}
    kb = await get_knowledge_base()
    query = q.strip().lower()
    results = []

    # Search FAQs
    for faq in (kb.get("faqs") or []):
        score = 0
        question = (faq.get("question") or "").lower()
        answer = (faq.get("answer") or "").lower()
        for word in query.split():
            if word in question:
                score += 2
            if word in answer:
                score += 1
        if score > 0:
            results.append({"section": "faqs", "title": faq.get("question", ""), "answer": faq.get("answer", ""), "score": score})

    # Search services
    for svc in (kb.get("services") or []):
        score = 0
        text = " ".join(str(svc.get(k, "")) for k in ("name", "category", "description")).lower()
        for word in query.split():
            if word in text:
                score += 2
        if score > 0:
            results.append({"section": "services", "title": svc.get("name", ""), "answer": svc.get("description", ""), "score": score})

    # Search packages
    for pkg in (kb.get("packages") or []):
        score = 0
        text = " ".join(str(pkg.get(k, "")) for k in ("package_name", "description", "included_items")).lower()
        for word in query.split():
            if word in text:
                score += 2
        if score > 0:
            results.append({"section": "packages", "title": pkg.get("package_name", ""), "answer": pkg.get("description", ""), "score": score})

    # Search about us / policies
    cp = kb.get("company_profile") or {}
    about = (cp.get("about_us") or "").lower()
    if about:
        score = sum(2 for w in query.split() if w in about)
        if score > 0:
            results.append({"section": "about_us", "title": "About Us", "answer": cp.get("about_us", ""), "score": score})

    pol = kb.get("policies") or {}
    for k, label in [("cancellation_policy", "Cancellation Policy"), ("refund_policy", "Refund Policy"),
                     ("appointment_policy", "Appointment Policy"), ("payment_policy", "Payment Policy")]:
        val = (pol.get(k) or "").lower()
        if val:
            score = sum(2 for w in query.split() if w in val)
            if score > 0:
                results.append({"section": "policies", "title": label, "answer": pol.get(k, ""), "score": score})

    # Search locations
    for loc in (kb.get("locations") or []):
        text = " ".join(str(loc.get(k, "")) for k in ("branch_name", "address", "city")).lower()
        score = sum(2 for w in query.split() if w in text)
        if score > 0:
            results.append({"section": "locations", "title": loc.get("branch_name", ""), "answer": f"{loc.get('address','')} {loc.get('city','')}", "score": score})

    results.sort(key=lambda x: x["score"], reverse=True)
    return {"query": q, "results": results[:10]}


@app.get("/api/knowledge-base/preview")
async def api_kb_preview():
    kb = await get_knowledge_base()
    context = build_knowledge_context(kb)
    return {"preview": context, "char_count": len(context)}


@app.post("/api/knowledge-base/sample")
async def api_kb_load_sample():
    saved = await save_knowledge_base(_KB_SAMPLE)
    return {"status": "sample_loaded", "sections": list(_KB_SAMPLE.keys())}


@app.get("/api/knowledge-base/{section}")
async def api_get_kb_section(section: str):
    if section not in _KB_SECTIONS:
        raise HTTPException(404, f"Unknown section: {section}. Valid: {_KB_SECTIONS}")
    data = await get_kb_section(section)
    return {"section": section, "data": data}


@app.post("/api/knowledge-base/{section}")
async def api_save_kb_section(section: str, req: Request):
    if section not in _KB_SECTIONS:
        raise HTTPException(400, f"Unknown section: {section}. Valid: {_KB_SECTIONS}")
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    await save_kb_section(section, data)
    return {"status": "saved", "section": section}


# ── WhatsApp Settings & Health ───────────────────────────────────────────────

@app.get("/api/whatsapp/settings")
async def api_get_wa_settings():
    return await get_wa_settings_masked()


@app.post("/api/whatsapp/settings")
async def api_save_wa_settings(req: WaSettingsRequest):
    if not isinstance(req.settings, dict):
        raise HTTPException(400, "settings must be a JSON object")
    await save_wa_settings(req.settings)
    return {"status": "saved", "keys_updated": [k for k in req.settings if k in WA_SETTINGS_KEYS]}


@app.get("/api/whatsapp/health")
async def api_wa_health():
    return await get_wa_health()


@app.post("/api/whatsapp/send-template")
async def api_wa_send_template(req: WaSendTemplateRequest):
    try:
        phone = normalize_phone(req.phone)
    except ValueError as e:
        raise HTTPException(400, str(e))
    result = await send_whatsapp_template(
        phone, req.template_name, req.language, req.parameters,
        event_type="manual_send", source_type="dashboard", source_id="manual",
    )
    return result


@app.get("/api/whatsapp/logs")
async def api_wa_logs(phone: Optional[str] = None, limit: int = 50):
    if phone:
        try:
            phone = normalize_phone(phone)
        except Exception:
            pass
    return {"logs": await get_whatsapp_logs(phone=phone, limit=min(limit, 200))}


# ── WhatsApp webhook stubs (Phase 8 — wired now so routes exist) ─────────────

@app.get("/api/whatsapp/webhook")
async def api_wa_webhook_verify(
    hub_mode: Optional[str] = Query(None, alias="hub.mode"),
    hub_verify_token: Optional[str] = Query(None, alias="hub.verify_token"),
    hub_challenge: Optional[str] = Query(None, alias="hub.challenge"),
):
    from db import get_setting as _gs
    verify_token = await _gs("WHATSAPP_VERIFY_TOKEN", "")
    if hub_mode == "subscribe" and verify_token and hub_verify_token == verify_token:
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(content=hub_challenge or "")
    raise HTTPException(403, "Verification failed")


@app.post("/api/whatsapp/webhook")
async def api_wa_webhook_receive(request: Request):
    try:
        payload = await request.json()
    except Exception:
        return {"status": "ok"}  # always return 200 to Meta
    try:
        parsed_list = parse_webhook_messages(payload)
        import asyncio as _asyncio
        for parsed in parsed_list:
            _asyncio.create_task(handle_inbound_whatsapp_message(parsed))
    except Exception as exc:
        logger.error("Webhook processing error: %s", exc)
    return {"status": "ok"}


# ── WhatsApp Inbox — Conversations ────────────────────────────────────────────

@app.get("/api/whatsapp/conversations")
async def api_wa_conversations(
    status: Optional[str] = None,
    ai_enabled: Optional[bool] = None,
    search: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    _user=Depends(require_auth),
):
    convs = await get_conversations(status=status, ai_enabled=ai_enabled, search=search, limit=limit, offset=offset)
    return {"conversations": convs, "total": len(convs)}


@app.get("/api/whatsapp/conversations/{conv_id}")
async def api_wa_conversation_detail(conv_id: str, _user=Depends(require_auth)):
    conv = await get_conversation_by_id(conv_id)
    if not conv:
        raise HTTPException(404, "Conversation not found")
    msgs = await get_messages(conv_id, limit=50)
    return {"conversation": conv, "messages": msgs}


@app.get("/api/whatsapp/conversations/{conv_id}/messages")
async def api_wa_conversation_messages(
    conv_id: str,
    limit: int = 50,
    offset: int = 0,
    _user=Depends(require_auth),
):
    msgs = await get_messages(conv_id, limit=limit, offset=offset)
    return {"messages": msgs}


class WaConvPatchRequest(BaseModel):
    status: Optional[str] = None
    ai_enabled: Optional[bool] = None
    assigned_to: Optional[str] = None
    unread_count: Optional[int] = None


@app.patch("/api/whatsapp/conversations/{conv_id}")
async def api_wa_patch_conversation(conv_id: str, req: WaConvPatchRequest, _user=Depends(require_auth)):
    updates = req.dict(exclude_none=True)
    if not updates:
        raise HTTPException(400, "No fields to update")
    conv = await patch_conversation(conv_id, updates)
    if not conv:
        raise HTTPException(404, "Conversation not found")
    return conv


class WaConvSendRequest(BaseModel):
    message: str


@app.post("/api/whatsapp/conversations/{conv_id}/send")
async def api_wa_conv_send(conv_id: str, req: WaConvSendRequest, _user=Depends(require_auth)):
    conv = await get_conversation_by_id(conv_id)
    if not conv:
        raise HTTPException(404, "Conversation not found")
    phone = conv.get("phone_number", "")
    message = (req.message or "").strip()
    if not message:
        raise HTTPException(400, "message is required")

    # Check 24h window
    window_open = await is_whatsapp_service_window_open(phone)
    if not window_open:
        raise HTTPException(400, {
            "message": "Customer service window may be closed. Please send an approved template.",
            "reason": "window_closed"
        })

    result = await send_whatsapp_text(phone, message)
    if not result.get("success"):
        err = result.get("error") or "Send failed"
        raise HTTPException(502, {"message": err, "reason": result.get("reason", "send_error")})

    provider_id = result.get("provider_message_id") or ""
    saved = await save_wa_message(
        conv_id=conv_id, phone=phone, direction="outbound",
        message_type="text", message_text=message,
        provider_message_id=provider_id, provider_status="sent",
        human_sent=True,
    )
    await update_conversation_last_message(conv_id, message, increment_unread=False)
    return {"status": "sent", "message": saved, "provider_message_id": provider_id}


class WaConvTemplateRequest(BaseModel):
    template_name: str
    language: Optional[str] = "en"
    parameters: Optional[list] = None


@app.post("/api/whatsapp/conversations/{conv_id}/send-template")
async def api_wa_conv_send_template(conv_id: str, req: WaConvTemplateRequest, _user=Depends(require_auth)):
    conv = await get_conversation_by_id(conv_id)
    if not conv:
        raise HTTPException(404, "Conversation not found")
    phone = conv.get("phone_number", "")
    result = await send_whatsapp_template(
        phone=phone,
        template_name=req.template_name,
        language=req.language or "en",
        parameters=req.parameters,
        event_type="manual_template",
        source_type="inbox",
        source_id=conv_id,
    )
    if not result.get("success"):
        err = result.get("error") or "Template send failed"
        raise HTTPException(502, {"message": err, "reason": result.get("reason", "send_error")})

    provider_id = result.get("provider_message_id") or ""
    saved = await save_wa_message(
        conv_id=conv_id, phone=phone, direction="outbound",
        message_type="template", message_text=f"[Template: {req.template_name}]",
        template_name=req.template_name,
        provider_message_id=provider_id, provider_status="sent",
        human_sent=True,
    )
    await update_conversation_last_message(conv_id, f"[Template: {req.template_name}]", increment_unread=False)
    return {"status": "sent", "message": saved, "provider_message_id": provider_id}


# ── Automation Rules ──────────────────────────────────────────────────────────

@app.get("/api/automation/rules")
async def api_get_automation_rules():
    return {"rules": await get_automation_rules(), "event_types": AUTOMATION_EVENT_TYPES, "action_types": AUTOMATION_ACTION_TYPES}


@app.post("/api/automation/rules")
async def api_save_automation_rules(req: AutomationRulesRequest):
    if not isinstance(req.rules, list):
        raise HTTPException(400, "rules must be a list")
    await save_automation_rules(req.rules)
    return {"status": "saved", "count": len(req.rules)}


@app.post("/api/automation/rules/reset")
async def api_reset_automation_rules():
    from whatsapp import _default_automation_rules
    default = _default_automation_rules()
    await save_automation_rules(default)
    return {"status": "reset", "count": len(default)}


@app.post("/api/automation/test")
async def api_test_automation(req: AutomationTestRequest):
    try:
        phone = normalize_phone(req.phone)
    except ValueError as e:
        raise HTTPException(400, str(e))
    rules = await get_automation_rules()
    rule = find_automation_rule(rules, req.event_type, req.source)
    if req.dry_run:
        return {
            "dry_run": True,
            "event_type": req.event_type,
            "source": req.source,
            "matched_rule": rule,
            "planned_action": rule.get("action") if rule else "no_matching_rule",
            "whatsapp_enabled": await get_wa_health(),
        }
    contact = {"phone": phone, "lead_name": req.lead_name or "Test Lead", "source": req.source or ""}
    result = await execute_automation_rule(req.event_type, contact)
    return {"dry_run": False, "event_type": req.event_type, **result}


# ── Automation Action Queue ───────────────────────────────────────────────────

@app.get("/api/automation/actions")
async def api_get_automation_actions(status: Optional[str] = None, phone: Optional[str] = None, limit: int = 100):
    if phone:
        try:
            phone = normalize_phone(phone)
        except Exception:
            pass
    actions = await get_automation_actions(status=status, phone=phone, limit=min(limit, 500))
    return {"actions": actions, "total": len(actions)}


@app.patch("/api/automation/actions/{action_id}/status")
async def api_update_action_status(action_id: str, req: StatusRequest):
    allowed = {"pending", "running", "completed", "failed", "cancelled", "waiting_schedule"}
    if req.status not in allowed:
        raise HTTPException(400, f"status must be one of: {', '.join(sorted(allowed))}")
    ok = await update_automation_action_status(action_id, req.status)
    if not ok:
        raise HTTPException(404, "Action not found")
    return {"status": "updated", "action_id": action_id}


@app.post("/api/automation/actions/run-due")
async def api_run_due_actions():
    result = await run_due_automation_actions()
    return {"status": "ok", **result}


# ── Lead Intake (n8n / Facebook / Website) ───────────────────────────────────

@app.post("/api/leads/incoming")
async def api_incoming_lead(req: LeadIncomingRequest):
    """Public endpoint for n8n, Facebook Lead Ads, website forms, etc."""
    try:
        phone = normalize_phone(req.phone)
    except ValueError as e:
        return JSONResponse(status_code=400, content={"success": False, "error": str(e)})

    source = (req.source or "api").strip().lower()
    lead_name = (req.name or "").strip() or "Unknown Lead"

    lead_data = {
        "phone_number": phone,
        "lead_name": lead_name,
        "email": req.email or "",
        "city": req.city or "",
        "source": source,
        "service_type": req.service or "",
        "business_name": req.business_name or "",
        "campaign_name": req.campaign_name or "",
        "crm_notes": req.message or "",
    }

    try:
        upsert_result = await upsert_crm_lead(lead_data, import_source=source)
        contact_status = upsert_result.get("status", "unknown")
    except Exception as exc:
        await log_error("leads_incoming", f"Upsert failed for {phone}", str(exc), "error")
        return JSONResponse(status_code=500, content={"success": False, "error": str(exc)})

    # Determine event type
    if contact_status == "duplicate":
        event_type = "re_enquiry"
    else:
        event_type = source_to_event_type(source)

    # Build contact dict for automation
    contact = {
        "phone": phone,
        "lead_name": lead_name,
        "business_name": req.business_name or "",
        "service_type": req.service or "",
        "source": source,
    }

    # Execute automation rule
    automation_result = await execute_automation_rule(event_type, contact)

    return {
        "success": True,
        "phone": phone,
        "contact_status": contact_status,
        "event_type": event_type,
        "automation_action": automation_result.get("action"),
        "automation_status": automation_result.get("automation_status"),
        "scheduled_action_id": automation_result.get("scheduled_action_id"),
        "whatsapp_status": automation_result.get("whatsapp_status"),
        "call_status": automation_result.get("call_status"),
    }


# ── Confirmation endpoints (Callback / Appointment / Showroom) ───────────────

@app.post("/api/whatsapp/confirm/callback")
async def api_wa_callback_confirm(req: WaConfirmRequest):
    try:
        phone = normalize_phone(req.phone)
    except ValueError as e:
        raise HTTPException(400, str(e))
    ctx = req.dict()
    ctx["phone"] = phone
    return await send_callback_confirmation(phone, ctx)


@app.post("/api/whatsapp/confirm/appointment")
async def api_wa_appointment_confirm(req: WaConfirmRequest):
    try:
        phone = normalize_phone(req.phone)
    except ValueError as e:
        raise HTTPException(400, str(e))
    ctx = req.dict()
    ctx["phone"] = phone
    return await send_appointment_confirmation(phone, ctx)


@app.post("/api/whatsapp/confirm/showroom")
async def api_wa_showroom_confirm(req: WaConfirmRequest):
    try:
        phone = normalize_phone(req.phone)
    except ValueError as e:
        raise HTTPException(400, str(e))
    ctx = req.dict()
    ctx["phone"] = phone
    return await send_showroom_visit_confirmation(phone, ctx)


@app.get("/api/settings")
async def api_get_settings():
    return await get_all_settings()


@app.post("/api/settings")
async def api_save_settings(req: SettingsRequest):
    filtered = {k: v for k, v in req.settings.items() if v is not None and v != ""}
    await save_settings(filtered)
    # VPS env vars are the single source of truth.
    # Only populate os.environ for keys that the host has NOT already defined.
    for k, v in filtered.items():
        if not os.environ.get(k):
            os.environ[k] = str(v)
    if any(k.startswith("RECORDING_") for k in filtered):
        await _schedule_recording_cleanup()
    return {"status": "saved", "count": len(filtered)}


@app.post("/api/setup/trunk")
async def api_setup_trunk():
    url = await eff("LIVEKIT_URL")
    key = await eff("LIVEKIT_API_KEY")
    secret = await eff("LIVEKIT_API_SECRET")
    sip_domain = await eff("VOBIZ_SIP_DOMAIN")
    username = await eff("VOBIZ_USERNAME")
    password = await eff("VOBIZ_PASSWORD")
    phone = await eff("VOBIZ_OUTBOUND_NUMBER")
    if not all([url, key, secret, sip_domain, username, password, phone]):
        raise HTTPException(400, "Configure LiveKit and Vobiz credentials in Settings first.")
    try:
        from livekit import api as lk_api
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        session = aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=ctx))
        lk = lk_api.LiveKitAPI(url=url, api_key=key, api_secret=secret, session=session)
        trunk = await lk.sip.create_sip_outbound_trunk(lk_api.CreateSIPOutboundTrunkRequest(trunk=lk_api.SIPOutboundTrunkInfo(name="Vobiz Outbound Trunk", address=sip_domain, auth_username=username, auth_password=password, numbers=[phone])))
        trunk_id = trunk.sip_trunk_id
        await set_setting("OUTBOUND_TRUNK_ID", trunk_id)
        if not os.environ.get("OUTBOUND_TRUNK_ID"):
            os.environ["OUTBOUND_TRUNK_ID"] = trunk_id
        await lk.aclose()
        await session.close()
        return {"status": "created", "trunk_id": trunk_id}
    except Exception as exc:
        raise HTTPException(500, f"Trunk creation failed: {exc}")


@app.get("/api/logs")
async def api_get_logs(limit: int = 200, level: Optional[str] = None, source: Optional[str] = None):
    return await get_logs(level=level, source=source, limit=limit)


@app.delete("/api/logs")
async def api_clear_logs():
    await clear_errors()
    return {"status": "cleared"}


def _clear_confirmation_error(req: Optional[ClearRecordsRequest]):
    if not req or req.confirm != "CLEAR_RECORDS":
        return JSONResponse(status_code=400, content={"success": False, "error": "Confirmation required"})
    return None


async def _admin_clear_response(req: Optional[ClearRecordsRequest], label: str, clear_fn):
    confirmation_error = _clear_confirmation_error(req)
    if confirmation_error:
        return confirmation_error
    deleted = await clear_fn()
    message = f"Cleared {deleted} {label}."
    logger.warning("Admin clear action: %s deleted=%s", label, deleted)
    if label != "error logs":
        await log_error("server", f"Admin clear action: {label}", f"deleted={deleted}", "warning")
    return {"success": True, "deleted": deleted, "message": message}


@app.delete("/api/admin/clear/call-logs")
async def api_admin_clear_call_logs(req: Optional[ClearRecordsRequest] = None):
    return await _admin_clear_response(req, "call logs", clear_call_logs)


@app.delete("/api/admin/clear/error-logs")
async def api_admin_clear_error_logs(req: Optional[ClearRecordsRequest] = None):
    return await _admin_clear_response(req, "error logs", clear_error_logs)


@app.delete("/api/admin/clear/contact-memory")
async def api_admin_clear_contact_memory(req: Optional[ClearRecordsRequest] = None):
    return await _admin_clear_response(req, "CRM memory records", clear_contact_memory)


@app.delete("/api/admin/clear/appointments")
async def api_admin_clear_appointments(req: Optional[ClearRecordsRequest] = None):
    return await _admin_clear_response(req, "appointments", clear_appointments)


@app.delete("/api/admin/clear/campaigns")
async def api_admin_clear_campaigns(req: Optional[ClearRecordsRequest] = None):
    return await _admin_clear_response(req, "campaigns", clear_campaigns)


@app.delete("/api/admin/clear/all-test-data")
async def api_admin_clear_all_test_data(req: Optional[ClearRecordsRequest] = None):
    confirmation_error = _clear_confirmation_error(req)
    if confirmation_error:
        return confirmation_error
    deleted_by_table = await clear_all_test_data()
    deleted = sum(deleted_by_table.values())
    logger.warning("Admin clear action: all test data deleted=%s details=%s", deleted, deleted_by_table)
    return {
        "success": True,
        "deleted": deleted,
        "message": "Cleared all test data. Settings and agent profiles were not changed.",
    }


@app.get("/api/crm")
async def api_get_contacts():
    return {"data": await get_contacts()}


@app.get("/api/crm/calls")
async def api_get_contact_calls(phone: str = Query(...)):
    return {"data": await get_calls_by_phone(phone)}


CRM_EXPORT_COLUMNS = [
    "lead_name", "phone_number", "email", "city", "location", "requirement", "budget",
    "source", "business_name", "campaign_name", "service_type", "crm_status",
    "custom_status", "next_followup_at", "assigned_to", "crm_notes",
    "last_call_outcome", "last_call_at", "total_calls", "created_at", "updated_at",
]


def _bool_query(value: Optional[str]):
    if value is None or value == "":
        return None
    return str(value).lower() in ("1", "true", "yes", "on")


def _crm_filter_dict(**kwargs) -> dict:
    return {k: v for k, v in kwargs.items() if v not in (None, "")}


async def _crm_contacts_from_filters(filters: dict) -> list:
    return await get_crm_contacts(
        status=filters.get("status"),
        outcome=filters.get("outcome"),
        q=filters.get("q"),
        due_today=_bool_query(filters.get("due_today")) is True,
        today=_bool_query(filters.get("today")) is True,
        timezone=filters.get("timezone"),
        source=filters.get("source"),
        business_name=filters.get("business_name"),
        campaign_name=filters.get("campaign_name"),
        city=filters.get("city"),
        date_from=filters.get("date_from"),
        date_to=filters.get("date_to"),
        assigned_to=filters.get("assigned_to"),
        recording_available=_bool_query(filters.get("recording_available")),
        has_followup=_bool_query(filters.get("has_followup")),
    )


# ── Outbound calling-window helpers ─────────────────────────────────────────

async def _outbound_tz() -> str:
    return (await eff("OUTBOUND_TIMEZONE") or "Asia/Kolkata").strip()


async def _outbound_window() -> tuple:
    """Return (start_hhmm, end_hhmm, allowed_day_abbrs) from env/settings."""
    tz = await _outbound_tz()
    start = (await eff("OUTBOUND_START_TIME") or "10:00").strip()
    end = (await eff("OUTBOUND_END_TIME") or "19:00").strip()
    days_raw = (await eff("OUTBOUND_ALLOWED_DAYS") or "mon,tue,wed,thu,fri,sat").strip().lower()
    days = {d.strip() for d in days_raw.split(",") if d.strip()}
    return tz, start, end, days


def _parse_hhmm(hhmm: str) -> tuple:
    """Parse '10:00' -> (10, 0). Safe fallback to (0, 0)."""
    try:
        h, m = hhmm.split(":")
        return int(h), int(m)
    except Exception:
        return 0, 0


async def _is_outbound_allowed() -> bool:
    """Return True if outbound calling is enabled and within the configured window."""
    enabled_raw = (await eff("OUTBOUND_CALLING_ENABLED") or "true").strip().lower()
    if enabled_raw in ("0", "false", "no", "off"):
        return False
    tz_name, start_str, end_str, allowed_days = await _outbound_window()
    try:
        from zoneinfo import ZoneInfo
        now_tz = datetime.now(tz=ZoneInfo(tz_name))
    except Exception:
        now_tz = datetime.now()
    day_abbr = now_tz.strftime("%a").lower()  # 'mon', 'tue', ...
    if day_abbr not in allowed_days:
        return False
    sh, sm = _parse_hhmm(start_str)
    eh, em = _parse_hhmm(end_str)
    current_minutes = now_tz.hour * 60 + now_tz.minute
    start_minutes = sh * 60 + sm
    end_minutes = eh * 60 + em
    return start_minutes <= current_minutes < end_minutes


async def _outbound_window_error() -> dict:
    """Build the structured 403 detail used by single-call and call-selected endpoints."""
    tz_name, start_str, end_str, allowed_days = await _outbound_window()
    # Compute next_allowed_at: next occurrence of start_str on an allowed day.
    try:
        from zoneinfo import ZoneInfo
        now_tz = datetime.now(tz=ZoneInfo(tz_name))
    except Exception:
        now_tz = datetime.now()
    sh, sm = _parse_hhmm(start_str)
    # Walk forward day-by-day until we land on an allowed day.
    from datetime import timedelta as _td
    candidate = now_tz.replace(hour=sh, minute=sm, second=0, microsecond=0)
    if candidate <= now_tz:
        candidate = candidate + _td(days=1)
    for _ in range(8):  # safety: max 7 days ahead
        if candidate.strftime("%a").lower() in allowed_days:
            break
        candidate = candidate + _td(days=1)
    try:
        next_allowed_at = candidate.isoformat()
    except Exception:
        next_allowed_at = None
    days_display = ", ".join(sorted(allowed_days))
    return {
        "error": "outside_outbound_window",
        "message": (
            f"Outbound calling is allowed only between {start_str} and {end_str} "
            f"{tz_name} ({days_display})."
        ),
        "start_time": start_str,
        "end_time": end_str,
        "timezone": tz_name,
        "allowed_days": sorted(allowed_days),
        "next_allowed_at": next_allowed_at,
    }


async def _wait_for_outbound_window(campaign_id: str, poll_seconds: int = 120) -> bool:
    """Block until the outbound calling window opens (or campaign is paused/stopped).

    Returns True if the window opened, False if the campaign was externally
    paused or stopped while waiting.
    """
    while not await _is_outbound_allowed():
        latest = await get_campaign(campaign_id)
        if not latest or (latest.get("status") in ("paused", "stopped", "completed")):
            return False
        tz_name, start_str, *_ = await _outbound_window()
        logger.info(
            "Campaign %s: outside outbound window (%s), sleeping %ds before retry",
            campaign_id, start_str, poll_seconds,
        )
        await asyncio.sleep(poll_seconds)
    return True


async def _import_leads(leads: list, import_source: str = "api", upload_batch_id: Optional[str] = None) -> dict:
    """Insert/update many leads, never aborting on a single bad row.

    The mobile number (normalized to E.164) is treated as the unique CRM
    identity. Existing phone numbers are merged into the existing record and
    counted as ``duplicate_count`` — *not* as ``imported_count``.
    """
    inserted = duplicates = skipped = failed = 0
    errors = []
    for idx, lead in enumerate(leads, start=1):
        if not isinstance(lead, dict):
            skipped += 1
            errors.append({"row": idx, "phone": "", "error": "row is not an object"})
            continue
        # Must have at least a phone before we even try.
        if not (lead.get("phone_number") or lead.get("phone")):
            skipped += 1
            errors.append({"row": idx, "phone": "", "error": "missing phone"})
            continue
        try:
            result = await upsert_crm_lead(lead, import_source=import_source, upload_batch_id=upload_batch_id)
            status = result.get("status")
            if status == "inserted":
                inserted += 1
            elif status == "duplicate":
                # Existing phone — merged + audit-noted, but reported separately.
                duplicates += 1
            else:
                # Defensive: any other status (e.g. legacy "updated") still
                # counts as a duplicate-merge for reporting purposes.
                duplicates += 1
        except Exception as exc:
            failed += 1
            errors.append({"row": idx, "phone": lead.get("phone_number") or lead.get("phone") or "", "error": str(exc)})
    return {
        "success": True,
        "total_rows": len(leads),
        "imported_count": inserted,
        "duplicate_count": duplicates,
        "updated_count": duplicates,  # alias kept for older UI builds
        "skipped_count": skipped,
        "failed_count": failed,
        # Back-compat keys (used by the original dashboard UI):
        "inserted": inserted,
        "updated": duplicates,
        "failed": failed,
        "errors": errors,
    }


UPLOAD_COLUMN_ALIASES = {
    "phone": "phone_number",
    "phone_number": "phone_number",
    "mobile": "phone_number",
    "mobile_number": "phone_number",
    "mobile_no": "phone_number",
    "contact": "phone_number",
    "contact_number": "phone_number",
    "name": "lead_name",
    "lead_name": "lead_name",
    "customer_name": "lead_name",
    "customer": "lead_name",
    "full_name": "lead_name",
    "email": "email",
    "email_id": "email",
    "city": "city",
    "town": "city",
    "location": "location",
    "area": "location",
    "source": "source",
    "lead_source": "source",
    "service": "service_type",
    "service_type": "service_type",
    "requirement": "service_type",
    "product": "service_type",
    "interest": "service_type",
    "notes": "crm_notes",
    "note": "crm_notes",
    "message": "crm_notes",
    "comment": "crm_notes",
    "comments": "crm_notes",
    "remark": "crm_notes",
    "remarks": "crm_notes",
    "status": "crm_status",
    "crm_status": "crm_status",
    "business_name": "business_name",
    "business": "business_name",
    "company": "business_name",
    "campaign": "campaign_name",
    "campaign_name": "campaign_name",
    "budget": "budget",
    "assigned_to": "assigned_to",
    "agent": "assigned_to",
    "followup": "next_followup_at",
    "follow_up": "next_followup_at",
    "next_followup_at": "next_followup_at",
}


def _normalize_upload_row(row: dict) -> dict:
    out = {}
    for key, value in (row or {}).items():
        if key is None:
            continue
        clean_key = str(key).strip().lower().replace(" ", "_").replace("-", "_")
        mapped = UPLOAD_COLUMN_ALIASES.get(clean_key, clean_key)
        if mapped in out and (value is None or str(value).strip() == ""):
            continue
        out[mapped] = value
    return out


@app.get("/api/lead-statuses")
async def api_get_lead_statuses():
    return {"data": await get_lead_statuses()}


@app.post("/api/lead-statuses")
async def api_add_lead_status(req: LeadStatusRequest):
    if not req.name.strip():
        raise HTTPException(400, "Status name is required")
    status = await add_lead_status(req.name, req.color)
    return {"success": True, "data": status}


@app.delete("/api/lead-statuses/{status_id}")
async def api_delete_lead_status(status_id: str):
    ok = await delete_lead_status(status_id)
    if not ok:
        raise HTTPException(404, "Status not found")
    return {"success": True, "message": "Status deleted"}


@app.post("/api/crm/contacts")
async def api_add_crm_contact(req: CrmLeadRequest):
    try:
        result = await upsert_crm_lead(
            req.dict(),
            import_source=req.source or "manual",
            forbid_duplicate=True,
        )
        # Trigger automation for manual lead
        phone = normalize_phone(req.phone_number or "")
        asyncio.create_task(execute_automation_rule(
            "manual_lead",
            {"phone": phone, "lead_name": req.lead_name or "", "source": req.source or "manual",
             "business_name": req.business_name or "", "service_type": req.service_type or ""},
        ))
        return {"success": True, **result, "message": "Lead added"}
    except DuplicateContactError as exc:
        # 409 Conflict so the Add-Lead UI shows a friendly error instead of
        # silently merging on top of an existing CRM record.
        raise HTTPException(
            status_code=409,
            detail={
                "error": "duplicate_contact",
                "phone_number": exc.phone,
                "message": "This mobile number already exists. Please open the existing lead and update it.",
            },
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc))


@app.post("/api/crm/import-leads")
async def api_import_crm_leads(req: CrmBulkImportRequest):
    return await _import_leads(req.leads or [], import_source=req.import_source or "api")


@app.post("/api/crm/upload-leads")
async def api_upload_crm_leads(file: UploadFile = File(...)):
    name = file.filename or ""
    content = await file.read()
    upload_batch_id = str(uuid.uuid4())
    rows: list = []
    try:
        if name.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")
            rows = [_normalize_upload_row(r) for r in csv.DictReader(StringIO(text))]
        elif name.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            if values:
                headers = [str(h or "").strip() for h in values[0]]
                rows = [_normalize_upload_row(dict(zip(headers, row))) for row in values[1:] if any(v is not None and str(v).strip() for v in row)]
        else:
            raise HTTPException(400, "Only .csv and .xlsx files are supported")

        if rows and not any(r.get("phone_number") or r.get("phone") for r in rows):
            raise HTTPException(400, "Required column missing: phone (aliases: phone_number, mobile, contact_number)")

        result = await _import_leads(rows, import_source="file_upload", upload_batch_id=upload_batch_id)
        # Trigger automation for each successfully inserted uploaded lead
        async def _trigger_upload_automation(rows_: list) -> None:
            for row_ in rows_:
                try:
                    ph_ = normalize_phone(row_.get("phone_number") or row_.get("phone") or "")
                    if ph_:
                        await execute_automation_rule(
                            "uploaded_lead",
                            {"phone": ph_, "lead_name": row_.get("lead_name") or "",
                             "source": row_.get("source") or "file_upload",
                             "business_name": row_.get("business_name") or "",
                             "service_type": row_.get("service_type") or ""},
                        )
                except Exception:
                    pass
        asyncio.create_task(_trigger_upload_automation(rows))
        return {**result, "upload_batch_id": upload_batch_id}
    except HTTPException:
        raise
    except Exception as exc:
        await log_error("server", "Lead upload failed", str(exc), "error")
        raise HTTPException(400, f"Upload failed: {exc}")


@app.get("/api/crm/sample-leads.csv")
async def api_sample_leads_csv():
    rows = [
        ["lead_name","phone_number","email","city","location","requirement","budget","source","business_name","campaign_name","service_type","crm_status","crm_notes","next_followup_at","assigned_to"],
        ["Ramesh","+919876543210","ramesh@gmail.com","Chennai","Tambaram","Villa plot","2500000","Facebook","Abhi Properties","Tambaram Villa Plot","Home visit","New","Interested in site visit","",""],
        ["Suresh","9876543211","suresh@gmail.com","Chennai","Velachery","Apartment","5000000","Website","Abhi Properties","Website Leads","Property consultation","New","Call after 5 PM","",""],
    ]
    out = StringIO()
    csv.writer(out).writerows(rows)
    out.seek(0)
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="sample_leads.csv"'})


@app.get("/api/crm/contacts")
async def api_get_crm_contacts(
    status: Optional[str] = None,
    outcome: Optional[str] = None,
    q: Optional[str] = None,
    due_today: bool = False,
    today: bool = False,
    timezone: Optional[str] = None,
    source: Optional[str] = None,
    business_name: Optional[str] = None,
    campaign_name: Optional[str] = None,
    city: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    assigned_to: Optional[str] = None,
    recording_available: Optional[str] = None,
    has_followup: Optional[str] = None,
):
    filters = _crm_filter_dict(
        status=status, outcome=outcome, q=q,
        due_today=str(due_today).lower() if due_today else "",
        today=str(today).lower() if today else "",
        timezone=timezone,
        source=source, business_name=business_name, campaign_name=campaign_name,
        city=city, date_from=date_from, date_to=date_to, assigned_to=assigned_to,
        recording_available=recording_available, has_followup=has_followup,
    )
    data = await _crm_contacts_from_filters(filters)
    return {"data": data, "total": len(data), "filters": filters}


@app.get("/api/crm/summary")
async def api_crm_summary():
    return await get_crm_summary()


@app.get("/api/crm/export/leads.csv")
async def api_export_crm_leads_csv(
    status: Optional[str] = None, outcome: Optional[str] = None, q: Optional[str] = None, due_today: Optional[str] = None,
    source: Optional[str] = None, business_name: Optional[str] = None, campaign_name: Optional[str] = None, city: Optional[str] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None, assigned_to: Optional[str] = None,
    recording_available: Optional[str] = None, has_followup: Optional[str] = None,
):
    filters = _crm_filter_dict(status=status, outcome=outcome, q=q, due_today=due_today, source=source, business_name=business_name, campaign_name=campaign_name, city=city, date_from=date_from, date_to=date_to, assigned_to=assigned_to, recording_available=recording_available, has_followup=has_followup)
    rows = await _crm_contacts_from_filters(filters)
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(CRM_EXPORT_COLUMNS)
    for row in rows:
        writer.writerow([row.get(c, "") for c in CRM_EXPORT_COLUMNS])
    out.seek(0)
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="crm_leads_export.csv"'})


@app.get("/api/crm/export/leads.xlsx")
async def api_export_crm_leads_xlsx(
    status: Optional[str] = None, outcome: Optional[str] = None, q: Optional[str] = None, due_today: Optional[str] = None,
    source: Optional[str] = None, business_name: Optional[str] = None, campaign_name: Optional[str] = None, city: Optional[str] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None, assigned_to: Optional[str] = None,
    recording_available: Optional[str] = None, has_followup: Optional[str] = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    filters = _crm_filter_dict(status=status, outcome=outcome, q=q, due_today=due_today, source=source, business_name=business_name, campaign_name=campaign_name, city=city, date_from=date_from, date_to=date_to, assigned_to=assigned_to, recording_available=recording_available, has_followup=has_followup)
    rows = await _crm_contacts_from_filters(filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "CRM Leads"
    ws.append(CRM_EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(c, "") for c in CRM_EXPORT_COLUMNS])
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 50)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="crm_leads_export.xlsx"'})


@app.get("/api/crm/contacts/{phone}")
async def api_get_crm_contact_detail(phone: str):
    detail = await get_crm_contact_detail(phone)
    if not detail.get("contact"):
        raise HTTPException(404, "CRM contact not found")
    return detail


@app.post("/api/crm/call-selected")
async def api_call_selected(req: CrmCallSelectedRequest):
    # ── Outbound calling-window guard ──────────────────────────────────────
    if not await _is_outbound_allowed():
        raise HTTPException(403, await _outbound_window_error())
    contacts = []
    failed = 0
    for phone in req.phones or []:
        try:
            clean = normalize_phone(phone)
            contact = await get_crm_contact_by_phone(clean)
            if not contact:
                failed += 1
                continue
            contacts.append({
                "phone": clean,
                "lead_name": contact.get("lead_name") or "there",
                "business_name": req.business_name or contact.get("business_name") or "our company",
                "service_type": req.service_type or contact.get("service_type") or "our service",
            })
        except Exception:
            failed += 1
    if not contacts:
        raise HTTPException(400, "No valid leads selected")
    campaign_id = await create_campaign(
        req.campaign_name or "Selected CRM Leads",
        json.dumps(contacts),
        "once",
        "09:00",
        req.call_delay_seconds or 15,
        req.system_prompt,
        req.agent_profile_id,
    )
    _start_campaign_task(campaign_id)
    return {"success": True, "dispatched": len(contacts), "failed": failed, "message": "Selected lead calling started"}


@app.put("/api/crm/contacts/{phone}")
async def api_update_crm_contact(phone: str, req: CrmLeadUpdateRequest):
    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No fields to update")
    ok = await update_crm_contact_full(phone, updates)
    if not ok:
        raise HTTPException(404, "CRM contact not found")
    return {"success": True, "message": "Lead updated", "phone_number": phone}


@app.patch("/api/crm/contacts/{phone}/status")
async def api_update_crm_status(phone: str, req: CrmStatusRequest):
    ok = await update_crm_contact_status(phone, req.crm_status, req.custom_status)
    if not ok:
        raise HTTPException(404, "CRM contact not found")
    return {"success": True, "message": "CRM status updated"}


@app.patch("/api/crm/contacts/{phone}/followup")
async def api_update_crm_followup(phone: str, req: CrmFollowupRequest):
    ok = await update_crm_contact_followup(phone, req.next_followup_at)
    if not ok:
        raise HTTPException(404, "CRM contact not found")
    return {"success": True, "message": "Follow-up updated"}


@app.patch("/api/crm/contacts/{phone}/notes")
async def api_update_crm_notes(phone: str, req: CrmNotesRequest):
    ok = await update_crm_contact_notes(phone, req.crm_notes)
    if not ok:
        raise HTTPException(404, "CRM contact not found")
    return {"success": True, "message": "CRM notes updated"}


@app.get("/api/agent-profiles")
async def api_list_agent_profiles():
    return await get_all_agent_profiles()


@app.post("/api/agent-profiles")
async def api_create_agent_profile(req: AgentProfileRequest):
    profile_id = await create_agent_profile(req.name, req.voice, req.model, req.system_prompt, req.enabled_tools, req.is_default)
    return {"status": "created", "id": profile_id}


@app.get("/api/agent-profiles/{profile_id}")
async def api_get_agent_profile(profile_id: str):
    profile = await get_agent_profile(profile_id)
    if not profile:
        raise HTTPException(404, "Profile not found")
    return profile


@app.put("/api/agent-profiles/{profile_id}")
async def api_update_agent_profile(profile_id: str, req: AgentProfileRequest):
    ok = await update_agent_profile(profile_id, {"name": req.name, "voice": req.voice, "model": req.model, "system_prompt": req.system_prompt, "enabled_tools": req.enabled_tools, "is_default": 1 if req.is_default else 0})
    if not ok:
        raise HTTPException(404, "Profile not found")
    return {"status": "updated"}


@app.delete("/api/agent-profiles/{profile_id}")
async def api_delete_agent_profile(profile_id: str):
    ok = await delete_agent_profile(profile_id)
    if not ok:
        raise HTTPException(404, "Profile not found")
    return {"status": "deleted"}


@app.post("/api/agent-profiles/{profile_id}/set-default")
async def api_set_default_profile(profile_id: str):
    await set_default_agent_profile(profile_id)
    return {"status": "default set"}


# ── Batch call in-memory store ──────────────────────────────────────
# Batches are ephemeral (process-lifetime). The per-item results are also
# written back to CRM contacts so nothing is lost if the process restarts.
_batch_store: "dict[str, dict]" = {}
_active_batch_tasks: "dict[str, asyncio.Task]" = {}

_BATCH_ITEM_OUTCOME_TO_CRM_STATUS = {
    "answered": "Contacted",
    "completed": "Contacted",
    "no_answer": "No Answer",
    "busy": "Busy",
    "failed": "Failed",
}


def _batch_summary(b: dict) -> dict:
    items = b.get("items", [])
    counts: dict = {s: 0 for s in ("pending", "calling", "answered", "no_answer", "busy", "failed", "completed", "cancelled")}
    current_phone = current_name = None
    for it in items:
        st = it.get("status", "pending")
        counts[st] = counts.get(st, 0) + 1
        if st == "calling":
            current_phone = it.get("phone")
            current_name = it.get("lead_name")
    return {
        "batch_id": b["batch_id"],
        "batch_name": b.get("batch_name", ""),
        "status": b.get("status", "pending"),
        "total_count": len(items),
        **counts,
        "current_phone": current_phone,
        "current_lead_name": current_name,
        "started_at": b.get("started_at"),
        "completed_at": b.get("completed_at"),
        "next_allowed_at": b.get("next_allowed_at"),
        "items": items,
    }


async def _run_batch(batch_id: str) -> None:
    b = _batch_store.get(batch_id)
    if not b:
        return
    b["status"] = "running"
    b["started_at"] = datetime.now().isoformat()
    items = b["items"]

    profile = None
    if b.get("agent_profile_id"):
        profile = await get_agent_profile(b["agent_profile_id"])

    url = await eff("LIVEKIT_URL")
    key = await eff("LIVEKIT_API_KEY")
    secret = await eff("LIVEKIT_API_SECRET")
    if not (url and key and secret):
        b["status"] = "failed"
        return

    delay_seconds = b.get("call_delay_seconds", 5)

    from livekit import api as lk_api_module
    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
    session = aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=ssl_ctx))
    try:
        lk = lk_api_module.LiveKitAPI(url=url, api_key=key, api_secret=secret, session=session)
        for idx, item in enumerate(items):
            if item.get("status") not in (None, "pending"):
                continue  # already done — skip on resume

            # Calling-window guard
            if not await _is_outbound_allowed():
                b["status"] = "waiting_schedule"
                err = await _outbound_window_error()
                b["next_allowed_at"] = err.get("next_allowed_at")
                opened = await _wait_for_batch_window(batch_id)
                if not opened:
                    break
                b["status"] = "running"
                b["next_allowed_at"] = None

            # Pause / stop check
            current = b.get("status")
            if current == "paused":
                break
            if current in ("stopped", "completed"):
                for it in items[idx:]:
                    if it.get("status") in (None, "pending"):
                        it["status"] = "cancelled"
                break

            phone = (item.get("phone") or "").strip()
            if not phone.startswith("+"):
                item["status"] = "failed"
                item["outcome"] = "invalid_phone"
                continue

            item["status"] = "calling"
            item["started_at"] = datetime.now().isoformat()

            room_name = f"batch-{batch_id[:8]}-{phone.replace('+','')}-{random.randint(100,999)}"
            item["room_name"] = room_name
            dispatched_at = time.time()
            contact = {
                "phone": phone,
                "lead_name": item.get("lead_name", "there"),
                "business_name": item.get("business_name", "our company"),
                "service_type": item.get("service_type", "our service"),
            }
            dispatched = await _dispatch_one(lk, lk_api_module, contact, room_name, b.get("system_prompt"), profile)
            if not dispatched:
                item["status"] = "failed"
                item["outcome"] = "dispatch_failed"
                item["ended_at"] = datetime.now().isoformat()
                continue

            # Sequential: wait for room to drain
            room_result = await _wait_for_room_finished(lk, lk_api_module, room_name)
            if room_result == "no_answer":
                item["status"] = "no_answer"
                item["outcome"] = "no_answer"
            elif room_result == "timeout":
                item["status"] = "failed"
                item["outcome"] = "timeout"
            else:
                outcome = await _classify_call_outcome(phone, dispatched_at, room_name)
                item["status"] = _BATCH_ITEM_OUTCOME_TO_CRM_STATUS.get(outcome, "completed").lower().replace(" ", "_")
                # Normalise to valid batch item status
                if item["status"] not in ("answered", "no_answer", "busy", "failed", "completed", "cancelled"):
                    item["status"] = "completed"
                item["outcome"] = outcome
            item["ended_at"] = datetime.now().isoformat()

            # Update CRM status for this lead
            try:
                crm_status = _BATCH_ITEM_OUTCOME_TO_CRM_STATUS.get(item.get("outcome", ""), "")
                if crm_status and phone:
                    await update_crm_contact_status(normalize_phone(phone), crm_status, None)
            except Exception as exc:
                logger.warning("Batch CRM status update failed for %s: %s", phone, exc)

            # WhatsApp fallback on call outcome
            try:
                call_outcome = item.get("outcome") or ""
                call_log_id = item.get("room_name") or ""
                contact_info = {"lead_name": item.get("lead_name", "there")}
                asyncio.create_task(handle_call_outcome_whatsapp_fallback(phone, call_outcome, call_log_id, contact_info))
            except Exception:
                pass

            # Check pause/stop after each call
            if b.get("status") in ("paused", "stopped"):
                break

            if idx < len(items) - 1 and delay_seconds:
                await asyncio.sleep(delay_seconds)

        await lk.aclose()
    except Exception as exc:
        logger.error("Batch run error: %s", exc)
        b["status"] = "failed"
    finally:
        await session.close()
        _active_batch_tasks.pop(batch_id, None)

    # Mark completed only when all items are terminal
    if b.get("status") not in ("paused", "stopped", "failed", "waiting_schedule"):
        all_done = all(it.get("status") not in (None, "pending", "calling") for it in items)
        b["status"] = "completed" if all_done else b.get("status", "completed")
    if b.get("status") == "completed":
        b["completed_at"] = datetime.now().isoformat()


async def _wait_for_batch_window(batch_id: str, poll_seconds: int = 120) -> bool:
    """Wait until outbound window opens, checking batch stop/pause between polls."""
    while not await _is_outbound_allowed():
        b = _batch_store.get(batch_id)
        if not b or b.get("status") in ("paused", "stopped", "completed"):
            return False
        tz_name, start_str, *_ = await _outbound_window()
        logger.info("Batch %s: outside outbound window (%s), sleeping %ds", batch_id, start_str, poll_seconds)
        await asyncio.sleep(poll_seconds)
    return True


def _start_batch_task(batch_id: str) -> None:
    existing = _active_batch_tasks.get(batch_id)
    if existing and not existing.done():
        return
    task = asyncio.create_task(_run_batch(batch_id))
    _active_batch_tasks[batch_id] = task


@app.post("/api/batch-calls")
async def api_create_batch(req: BatchCallRequest):
    if not await _is_outbound_allowed():
        raise HTTPException(403, await _outbound_window_error())
    if not req.contacts:
        raise HTTPException(400, "contacts list cannot be empty")
    batch_id = str(uuid.uuid4())
    items = []
    seen: set = set()
    for c in req.contacts:
        phone = (c.get("phone") or "").strip()
        if not phone or phone in seen:
            continue
        seen.add(phone)
        items.append({
            "phone": phone,
            "lead_name": c.get("lead_name") or c.get("name") or "there",
            "business_name": c.get("business_name") or "our company",
            "service_type": c.get("service_type") or "our service",
            "status": "pending",
            "outcome": None,
            "room_name": None,
            "started_at": None,
            "ended_at": None,
        })
    if not items:
        raise HTTPException(400, "No valid contacts after deduplication")
    b = {
        "batch_id": batch_id,
        "batch_name": req.batch_name or f"Batch {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "status": "pending",
        "items": items,
        "call_delay_seconds": req.call_delay_seconds,
        "agent_profile_id": req.agent_profile_id,
        "system_prompt": req.system_prompt,
        "started_at": None,
        "completed_at": None,
        "next_allowed_at": None,
    }
    _batch_store[batch_id] = b
    _start_batch_task(batch_id)
    return _batch_summary(b)


@app.get("/api/batch-calls")
async def api_list_batches():
    return [_batch_summary(b) for b in _batch_store.values()]


@app.get("/api/batch-calls/{batch_id}")
async def api_get_batch(batch_id: str):
    b = _batch_store.get(batch_id)
    if not b:
        raise HTTPException(404, "Batch not found")
    return _batch_summary(b)


@app.patch("/api/batch-calls/{batch_id}/status")
async def api_update_batch_status(batch_id: str, req: StatusRequest):
    b = _batch_store.get(batch_id)
    if not b:
        raise HTTPException(404, "Batch not found")
    allowed = {"paused", "running", "stopped"}
    if req.status not in allowed:
        raise HTTPException(400, f"status must be one of: {', '.join(sorted(allowed))}")
    if req.status == "stopped":
        b["status"] = "stopped"
        for it in b["items"]:
            if it.get("status") in (None, "pending"):
                it["status"] = "cancelled"
    elif req.status == "paused":
        b["status"] = "paused"
    elif req.status == "running":
        b["status"] = "running"
        _start_batch_task(batch_id)  # resumes — runner skips non-pending items
    return _batch_summary(b)


# ── Campaign sequencing ─────────────────────────────────────────────
# These statuses live in each ``contact`` dict inside ``contacts_json`` and are
# persisted back to Supabase after every transition so pause/resume/stop
# survives a process restart.
CAMPAIGN_TERMINAL_LEAD_STATUSES = {
    "completed", "answered", "no_answer", "busy", "failed", "skipped", "cancelled",
}

# Map call_logs.outcome → campaign lead status.
_OUTCOME_TO_LEAD_STATUS = {
    "booked": "answered",
    "answered": "answered",
    "interested": "answered",
    "not_interested": "answered",
    "callback_requested": "answered",
    "transferred": "answered",
    "completed": "completed",
    "no_answer": "no_answer",
    "busy": "busy",
    "failed": "failed",
    "voicemail": "no_answer",
}

# Tasks currently executing each campaign (so /run can be idempotent).
_active_campaign_tasks: "dict[str, asyncio.Task]" = {}


def _campaign_env_int(name: str, default: int) -> int:
    try:
        return max(int(os.getenv(name, str(default))), 0)
    except (TypeError, ValueError):
        return default


async def _campaign_max_concurrent() -> int:
    raw = await eff("MAX_CONCURRENT_OUTBOUND_CALLS")
    try:
        return max(int(raw or "1"), 1)
    except (TypeError, ValueError):
        return 1


async def _campaign_call_delay(default: int) -> int:
    raw = await eff("OUTBOUND_CALL_DELAY_SECONDS")
    try:
        return max(int(raw or str(default)), 0)
    except (TypeError, ValueError):
        return default


def _norm_lead_status(value: Optional[str]) -> str:
    return (value or "").strip().lower() or "pending"


async def _build_dispatch_metadata(contact: dict, prompt: Optional[str], profile: Optional[dict]) -> dict:
    lead_name = contact.get("lead_name", "there")
    business_name = contact.get("business_name", "our company")
    service_type = contact.get("service_type", "our service")

    # Prompt priority: explicit > agent profile > legacy global > call-type resolved
    saved_prompt = prompt or (await get_setting("system_prompt", "")) or None
    if profile and not saved_prompt and profile.get("system_prompt"):
        saved_prompt = profile["system_prompt"]
    if not saved_prompt:
        call_type = contact.get("call_type") or "welcome_call"
        saved_prompt = await resolve_ai_prompt(
            call_type=call_type,
            lead_name=lead_name,
            business_name=business_name,
            service_type=service_type,
        )

    metadata = {
        "phone_number": contact["phone"],
        "lead_name": lead_name,
        "business_name": business_name,
        "service_type": service_type,
        "system_prompt": saved_prompt,
    }
    if profile:
        if profile.get("voice"):
            metadata["voice_override"] = profile["voice"]
        if profile.get("model"):
            metadata["model_override"] = profile["model"]
        if profile.get("enabled_tools"):
            metadata["tools_override"] = profile["enabled_tools"]
    return metadata


async def _dispatch_one(lk, lk_api, contact: dict, room_name: str, prompt: Optional[str], profile: Optional[dict] = None) -> bool:
    try:
        await lk.room.create_room(lk_api.CreateRoomRequest(name=room_name, empty_timeout=300, max_participants=5))
        metadata = await _build_dispatch_metadata(contact, prompt, profile)
        await lk.agent_dispatch.create_dispatch(lk_api.CreateAgentDispatchRequest(agent_name="outbound-caller", room=room_name, metadata=json.dumps(metadata)))
        return True
    except Exception as exc:
        logger.error("Campaign dispatch error for %s: %s", contact.get("phone"), exc)
        return False


async def _wait_for_room_finished(lk, lk_api_module, room_name: str, *, max_wait: int = 3600, poll_interval: int = 8) -> str:
    """Block until the dispatched LiveKit room is empty (call ended).

    Returns one of: ``completed`` (room gone / both peers left), ``no_answer``
    (SIP never joined within the answer-grace window), ``timeout`` (safety cap).
    This is what gives us *sequential* calling: we don't start lead N+1 until
    lead N's room has shut down.
    """
    start = time.time()
    answer_grace = start + 90  # SIP has 90s to actually ring + answer
    sip_seen = False
    while True:
        if time.time() - start > max_wait:
            return "timeout"
        try:
            resp = await lk.room.list_rooms(lk_api_module.ListRoomsRequest(names=[room_name]))
            rooms = list(getattr(resp, "rooms", []) or [])
        except Exception as exc:
            logger.debug("list_rooms error for %s: %s", room_name, exc)
            await asyncio.sleep(poll_interval)
            continue
        if not rooms:
            # Room is gone — either never created or already torn down.
            if sip_seen or time.time() > answer_grace:
                return "completed"
            await asyncio.sleep(poll_interval)
            continue
        room = rooms[0]
        num_participants = int(getattr(room, "num_participants", 0) or 0)
        if num_participants >= 2:
            sip_seen = True
        if sip_seen and num_participants <= 1:
            # Caller hung up: only the agent (or no one) is left.
            return "completed"
        if not sip_seen and time.time() > answer_grace:
            return "no_answer"
        await asyncio.sleep(poll_interval)


async def _classify_call_outcome(phone: str, dispatch_time: float, room_name: str) -> str:
    """After the room closes, look up the call_log entry the agent just wrote."""
    try:
        calls = await get_calls_by_phone(phone)
    except Exception:
        return "completed"
    for call in calls or []:
        ts = call.get("timestamp") or ""
        try:
            # call_logs.timestamp is ISO; compare as string within the last hour.
            if ts and ts >= datetime.fromtimestamp(dispatch_time).isoformat()[:19]:
                outcome = (call.get("outcome") or "").strip().lower()
                return _OUTCOME_TO_LEAD_STATUS.get(outcome, "completed")
        except Exception:
            continue
    return "completed"


async def _run_campaign(campaign_id: str) -> None:
    campaign = await get_campaign(campaign_id)
    if not campaign:
        return
    try:
        contacts = json.loads(campaign.get("contacts_json") or "[]")
    except Exception:
        contacts = []
    if not isinstance(contacts, list) or not contacts:
        return

    # De-dupe by phone within this campaign so the same number can't be queued
    # twice in a single run.
    seen_phones: set = set()
    deduped: list = []
    for contact in contacts:
        if not isinstance(contact, dict):
            continue
        phone = (contact.get("phone") or "").strip()
        if not phone:
            continue
        contact.setdefault("status", "pending")
        if phone in seen_phones and _norm_lead_status(contact.get("status")) == "pending":
            contact["status"] = "skipped"
            contact["skip_reason"] = "duplicate phone in campaign"
        seen_phones.add(phone)
        deduped.append(contact)
    contacts = deduped
    await update_campaign_contacts(campaign_id, contacts)

    profile = await get_agent_profile(campaign.get("agent_profile_id")) if campaign.get("agent_profile_id") else None
    url = await eff("LIVEKIT_URL")
    key = await eff("LIVEKIT_API_KEY")
    secret = await eff("LIVEKIT_API_SECRET")
    if not (url and key and secret):
        logger.error("Campaign %s: LiveKit not configured", campaign_id)
        return

    max_concurrent = await _campaign_max_concurrent()
    delay_seconds = await _campaign_call_delay(int(campaign.get("call_delay_seconds") or 30))

    from livekit import api as lk_api_module
    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
    session = aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=ssl_ctx))
    ok_count = fail_count = 0
    final_status = "completed"
    try:
        lk = lk_api_module.LiveKitAPI(url=url, api_key=key, api_secret=secret, session=session)
        for idx, contact in enumerate(contacts):
            # Re-read campaign status on every iteration so pause/stop signals
            # arrive at the next safe boundary.
            latest = await get_campaign(campaign_id)
            current_status = (latest or {}).get("status") if latest else None
            if current_status == "paused":
                logger.info("Campaign %s paused — stopping after current safe boundary", campaign_id)
                final_status = "paused"
                break
            if current_status in ("stopped", "completed"):
                logger.info("Campaign %s %s — cancelling pending leads", campaign_id, current_status)
                final_status = current_status
                # Mark every still-pending lead as cancelled so resume can't run them.
                for c in contacts[idx:]:
                    if _norm_lead_status(c.get("status")) == "pending":
                        c["status"] = "cancelled"
                break

            if _norm_lead_status(contact.get("status")) != "pending":
                continue  # already handled in a previous run (resume case)

            # ── Outbound calling-window guard ──────────────────────────────
            # If we're currently outside the allowed calling window, wait until
            # it opens (or the campaign is paused/stopped externally).  Leads
            # are NEVER skipped or cancelled just because of the time — they
            # remain pending until we get a window.
            if not await _is_outbound_allowed():
                opened = await _wait_for_outbound_window(campaign_id)
                if not opened:
                    # Campaign was paused/stopped while waiting — exit cleanly.
                    latest = await get_campaign(campaign_id)
                    final_status = (latest or {}).get("status") or "paused"
                    break
                # Re-check campaign status now that time passed.
                latest = await get_campaign(campaign_id)
                current_status = (latest or {}).get("status") if latest else None
                if current_status in ("paused", "stopped", "completed"):
                    final_status = current_status
                    break

            phone = (contact.get("phone") or "").strip()
            if not phone.startswith("+"):
                contact["status"] = "failed"
                contact["failure_reason"] = "phone not in E.164 format"
                fail_count += 1
                await update_campaign_contacts(campaign_id, contacts)
                continue

            contact["status"] = "in_progress"
            contact["started_at"] = datetime.now().isoformat()
            await update_campaign_contacts(campaign_id, contacts)

            room_name = f"camp-{campaign_id[:8]}-{phone.replace('+','')}-{random.randint(100,999)}"
            contact["room_name"] = room_name
            dispatched_at = time.time()
            dispatched = await _dispatch_one(lk, lk_api_module, contact, room_name, campaign.get("system_prompt"), profile)
            if not dispatched:
                contact["status"] = "failed"
                contact["failure_reason"] = "dispatch failed"
                fail_count += 1
                await update_campaign_contacts(campaign_id, contacts)
                continue

            # Sequential mode: wait for the room to drain before moving on.
            if max_concurrent <= 1:
                room_result = await _wait_for_room_finished(lk, lk_api_module, room_name)
                if room_result == "no_answer":
                    contact["status"] = "no_answer"
                elif room_result == "timeout":
                    contact["status"] = "failed"
                    contact["failure_reason"] = "safety timeout while waiting for room"
                else:
                    contact["status"] = await _classify_call_outcome(phone, dispatched_at, room_name)
            else:
                # Parallel mode: trust the agent to write a call_log and move on
                # after a short stagger.
                contact["status"] = "completed"

            contact["finished_at"] = datetime.now().isoformat()
            ok_count += 1 if contact["status"] in ("answered", "completed") else 0
            if contact["status"] in ("failed", "no_answer", "busy"):
                fail_count += 1
            await update_campaign_contacts(campaign_id, contacts)

            # WhatsApp fallback on campaign call outcome
            try:
                call_outcome = contact.get("status") or ""
                call_log_id = contact.get("room_name") or ""
                contact_info = {"lead_name": contact.get("lead_name", "there")}
                asyncio.create_task(handle_call_outcome_whatsapp_fallback(phone, call_outcome, call_log_id, contact_info))
            except Exception:
                pass

            # Pause/Stop can come in *during* a call — re-check right after
            # finishing this one and don't sleep into the next dispatch.
            latest = await get_campaign(campaign_id)
            current_status = (latest or {}).get("status") if latest else None
            if current_status == "paused":
                final_status = "paused"
                break
            if current_status in ("stopped", "completed"):
                final_status = current_status
                break

            # Inter-call delay (only between calls, not after the last one).
            if idx < len(contacts) - 1 and delay_seconds:
                await asyncio.sleep(delay_seconds)
        await lk.aclose()
    except Exception as exc:
        logger.error("Campaign run error: %s", exc)
        await log_error("server", f"Campaign {campaign_id} run error", str(exc), "error")
    finally:
        await session.close()
        _active_campaign_tasks.pop(campaign_id, None)

    # If everything got through, mark completed; otherwise leave whatever
    # pause/stop state we detected.
    if final_status == "completed" and all(_norm_lead_status(c.get("status")) in CAMPAIGN_TERMINAL_LEAD_STATUSES for c in contacts):
        await update_campaign_run_stats(campaign_id, ok_count, fail_count, status="completed")
    else:
        await update_campaign_run_stats(campaign_id, ok_count, fail_count, status=final_status)


def _start_campaign_task(campaign_id: str) -> None:
    existing = _active_campaign_tasks.get(campaign_id)
    if existing and not existing.done():
        return  # already running, don't double-spawn
    task = asyncio.create_task(_run_campaign(campaign_id))
    _active_campaign_tasks[campaign_id] = task


async def _reschedule_all_campaigns() -> None:
    if not _scheduler:
        return
    try:
        for c in await get_all_campaigns():
            if c.get("status") == "active" and c.get("schedule_type") in ("daily", "weekdays"):
                _schedule_campaign(c["id"], c["schedule_type"], c.get("schedule_time", "09:00"))
    except Exception as exc:
        logger.warning("Could not reschedule campaigns: %s", exc)


async def _scheduled_recording_cleanup() -> None:
    try:
        if await _recording_auto_delete_enabled():
            await _cleanup_old_recordings()
    except Exception as exc:
        logger.exception("Scheduled recording cleanup failed: %s", exc)
        await log_error("server", "Scheduled recording cleanup failed", str(exc), "error")


async def _schedule_recording_cleanup() -> None:
    if not _scheduler:
        return
    job_id = "recording_cleanup"
    if _scheduler.get_job(job_id):
        _scheduler.remove_job(job_id)
    if not await _recording_auto_delete_enabled():
        return
    cleanup_time = await _recording_cleanup_time()
    try:
        hour, minute = map(int, cleanup_time.split(":"))
    except (ValueError, AttributeError):
        hour, minute = 2, 0
    trigger = CronTrigger(hour=hour, minute=minute, timezone=ZoneInfo("Asia/Kolkata"))
    _scheduler.add_job(_scheduled_recording_cleanup, trigger=trigger, id=job_id, replace_existing=True)


def _schedule_campaign(campaign_id: str, schedule_type: str, schedule_time: str) -> None:
    if not _scheduler:
        return
    job_id = f"campaign_{campaign_id}"
    if _scheduler.get_job(job_id):
        _scheduler.remove_job(job_id)
    try:
        hour, minute = map(int, schedule_time.split(":"))
    except (ValueError, AttributeError):
        hour, minute = 9, 0
    trigger = CronTrigger(hour=hour, minute=minute) if schedule_type == "daily" else CronTrigger(day_of_week="mon-fri", hour=hour, minute=minute)
    _scheduler.add_job(_run_campaign, trigger=trigger, args=[campaign_id], id=job_id, replace_existing=True)


@app.post("/api/campaigns")
async def api_create_campaign(req: CampaignRequest):
    if not req.contacts:
        raise HTTPException(400, "contacts list cannot be empty")
    if req.schedule_type not in ("once", "daily", "weekdays"):
        raise HTTPException(400, "schedule_type must be: once | daily | weekdays")
    campaign_id = await create_campaign(req.name, json.dumps(req.contacts), req.schedule_type, req.schedule_time, req.call_delay_seconds, req.system_prompt, req.agent_profile_id)
    campaign = await get_campaign(campaign_id)
    if req.schedule_type == "once":
        _start_campaign_task(campaign_id)
    else:
        _schedule_campaign(campaign_id, req.schedule_type, req.schedule_time)
    return {"status": "created", "campaign_id": campaign_id, "campaign": campaign}


@app.get("/api/campaigns")
async def api_list_campaigns():
    return await get_all_campaigns()


@app.delete("/api/campaigns/{campaign_id}")
async def api_delete_campaign(campaign_id: str):
    ok = await delete_campaign(campaign_id)
    if not ok:
        raise HTTPException(404, "Campaign not found")
    job_id = f"campaign_{campaign_id}"
    if _scheduler and _scheduler.get_job(job_id):
        _scheduler.remove_job(job_id)
    return {"status": "deleted"}


@app.post("/api/campaigns/{campaign_id}/run")
async def api_run_campaign_now(campaign_id: str):
    if not await get_campaign(campaign_id):
        raise HTTPException(404, "Campaign not found")
    await update_campaign_status(campaign_id, "active")
    _start_campaign_task(campaign_id)
    return {"status": "dispatching", "campaign_id": campaign_id}


@app.get("/api/campaigns/{campaign_id}/progress")
async def api_campaign_progress(campaign_id: str):
    campaign = await get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(404, "Campaign not found")
    try:
        contacts = json.loads(campaign.get("contacts_json") or "[]")
    except Exception:
        contacts = []
    counts: dict = {s: 0 for s in ("pending", "in_progress", "calling", "completed", "answered", "no_answer", "busy", "failed", "skipped", "cancelled")}
    current_phone = current_name = None
    for c in contacts:
        st = _norm_lead_status(c.get("status"))
        counts[st] = counts.get(st, 0) + 1
        if st in ("in_progress", "calling"):
            current_phone = c.get("phone")
            current_name = c.get("lead_name")
    return {
        "campaign_id": campaign_id,
        "name": campaign.get("name"),
        "status": campaign.get("status"),
        "total_leads": len(contacts),
        "total_count": len(contacts),
        "current_phone": current_phone,
        "current_lead_name": current_name,
        "counts": counts,
        **counts,
        "leads": [
            {
                "phone": c.get("phone"),
                "lead_name": c.get("lead_name"),
                "status": _norm_lead_status(c.get("status")),
                "started_at": c.get("started_at"),
                "finished_at": c.get("finished_at"),
                "room_name": c.get("room_name"),
                "failure_reason": c.get("failure_reason") or c.get("skip_reason"),
            } for c in contacts
        ],
    }


@app.patch("/api/campaigns/{campaign_id}/status")
async def api_update_campaign_status(campaign_id: str, req: StatusRequest):
    allowed = {"active", "paused", "completed", "stopped"}
    if req.status not in allowed:
        raise HTTPException(400, f"status must be one of: {', '.join(sorted(allowed))}")
    campaign = await get_campaign(campaign_id)
    if not campaign:
        raise HTTPException(404, "Campaign not found")

    ok = await update_campaign_status(campaign_id, req.status)
    if not ok:
        raise HTTPException(404, "Campaign not found")

    job_id = f"campaign_{campaign_id}"
    if req.status in ("paused", "stopped", "completed") and _scheduler and _scheduler.get_job(job_id):
        _scheduler.remove_job(job_id)

    if req.status == "stopped":
        # Mark every still-pending lead cancelled so resume cannot run them.
        try:
            contacts = json.loads(campaign.get("contacts_json") or "[]")
        except Exception:
            contacts = []
        changed = False
        for c in contacts:
            if _norm_lead_status(c.get("status")) == "pending":
                c["status"] = "cancelled"
                changed = True
        if changed:
            await update_campaign_contacts(campaign_id, contacts)

    if req.status == "active":
        # Resume: only the still-pending leads will be picked up (terminal
        # statuses are skipped inside the loop), so no number is ever called twice.
        if campaign.get("schedule_type") in ("daily", "weekdays"):
            _schedule_campaign(campaign_id, campaign["schedule_type"], campaign.get("schedule_time", "09:00"))
        else:
            _start_campaign_task(campaign_id)

    return {"status": req.status}


# ── Scheduled Calls (Due Today) ─────────────────────────────────────────

@app.get("/api/scheduled-calls")
async def api_scheduled_calls(timezone: Optional[str] = None):
    """Return CRM leads grouped by their scheduled follow-up state."""
    tz = timezone or (await eff("OUTBOUND_TIMEZONE")) or "Asia/Kolkata"
    today_str = _tz_today(tz)
    all_leads = await get_crm_contacts(timezone=tz)
    window_ok = await _is_outbound_allowed()
    win_err = None if window_ok else await _outbound_window_error()

    due_today, upcoming, overdue, completed_today = [], [], [], []
    for lead in all_leads:
        status = (lead.get("crm_status") or "New")
        if status in CRM_TERMINAL_STATUSES:
            continue
        fup = (lead.get("next_followup_at") or "")[:10]
        if not fup:
            continue
        if fup == today_str:
            due_today.append(lead)
        elif fup < today_str:
            overdue.append(lead)
        elif fup > today_str:
            upcoming.append(lead)

    return {
        "window_allowed": window_ok,
        "window_error": win_err,
        "today": today_str,
        "timezone": tz,
        "due_today": due_today,
        "overdue": overdue,
        "upcoming": upcoming,
        "completed_today": completed_today,
        "counts": {
            "due_today": len(due_today),
            "overdue": len(overdue),
            "upcoming": len(upcoming),
        },
    }


@app.post("/api/scheduled-calls/start-due-today")
async def api_start_due_today(timezone: Optional[str] = Query(None)):
    """Create + start a batch from today’s Due Today CRM leads."""
    if not await _is_outbound_allowed():
        raise HTTPException(403, await _outbound_window_error())
    tz = timezone or (await eff("OUTBOUND_TIMEZONE")) or "Asia/Kolkata"
    today_str = _tz_today(tz)
    all_leads = await get_crm_contacts(due_today=True, timezone=tz)
    if not all_leads:
        raise HTTPException(404, "No Due Today leads found")
    contacts = [
        {
            "phone": lead["phone_number"],
            "lead_name": lead.get("lead_name") or "there",
            "business_name": lead.get("business_name") or "our company",
            "service_type": lead.get("service_type") or "our service",
            "call_type": _crm_status_to_call_type(lead.get("crm_status") or ""),
        }
        for lead in all_leads
        if lead.get("phone_number")
    ]
    batch_name = f"Due Today Calls - {today_str}"
    # Create via the existing batch machinery
    batch_id = str(uuid.uuid4())
    items = []
    seen: set = set()
    for c in contacts:
        phone = (c.get("phone") or "").strip()
        if not phone or phone in seen:
            continue
        seen.add(phone)
        items.append({
            "phone": phone,
            "lead_name": c.get("lead_name", "there"),
            "business_name": c.get("business_name", "our company"),
            "service_type": c.get("service_type", "our service"),
            "status": "pending", "outcome": None, "room_name": None,
            "started_at": None, "ended_at": None,
        })
    if not items:
        raise HTTPException(404, "No valid leads to call")
    b = {
        "batch_id": batch_id, "batch_name": batch_name, "status": "pending",
        "items": items, "call_delay_seconds": 10, "agent_profile_id": None,
        "system_prompt": None, "started_at": None, "completed_at": None, "next_allowed_at": None,
    }
    _batch_store[batch_id] = b
    _start_batch_task(batch_id)
    return {"batch_id": batch_id, "batch_name": batch_name, "total": len(items), **_batch_summary(b)}


# ── Outbound Schedule Settings ─────────────────────────────────────────

_SCHEDULE_KEYS = (
    "OUTBOUND_TIMEZONE", "OUTBOUND_START_TIME", "OUTBOUND_END_TIME",
    "OUTBOUND_ALLOWED_DAYS", "OUTBOUND_CALLING_ENABLED",
)


@app.get("/api/outbound/schedule/settings")
async def api_get_outbound_schedule_settings():
    result = {}
    for key in _SCHEDULE_KEYS:
        result[key] = await eff(key)
    # Apply defaults for missing values
    result.setdefault("OUTBOUND_TIMEZONE", "Asia/Kolkata")
    if not result["OUTBOUND_TIMEZONE"]:
        result["OUTBOUND_TIMEZONE"] = "Asia/Kolkata"
    result.setdefault("OUTBOUND_START_TIME", "10:00")
    if not result["OUTBOUND_START_TIME"]:
        result["OUTBOUND_START_TIME"] = "10:00"
    result.setdefault("OUTBOUND_END_TIME", "19:00")
    if not result["OUTBOUND_END_TIME"]:
        result["OUTBOUND_END_TIME"] = "19:00"
    result.setdefault("OUTBOUND_ALLOWED_DAYS", "mon,tue,wed,thu,fri,sat")
    if not result["OUTBOUND_ALLOWED_DAYS"]:
        result["OUTBOUND_ALLOWED_DAYS"] = "mon,tue,wed,thu,fri,sat"
    result.setdefault("OUTBOUND_CALLING_ENABLED", "true")
    if not result["OUTBOUND_CALLING_ENABLED"]:
        result["OUTBOUND_CALLING_ENABLED"] = "true"
    return result


@app.post("/api/outbound/schedule/settings")
async def api_save_outbound_schedule_settings(req: OutboundScheduleRequest):
    data = {k: v for k, v in req.dict().items() if v is not None and v != ""}
    if not data:
        raise HTTPException(400, "No settings provided")
    await save_settings(data)
    # Propagate to os.environ so the running process picks them up immediately
    # without needing a restart (only fills gaps, honours existing VPS env vars).
    for k, v in data.items():
        if not os.environ.get(k):
            os.environ[k] = v
    # Return the updated schedule status
    return await api_get_outbound_schedule_status()


@app.get("/api/outbound/schedule/status")
async def api_get_outbound_schedule_status():
    tz_name, start_str, end_str, allowed_days = await _outbound_window()
    enabled_raw = (await eff("OUTBOUND_CALLING_ENABLED") or "true").strip().lower()
    calling_enabled = enabled_raw not in ("0", "false", "no", "off")
    allowed = calling_enabled and await _is_outbound_allowed()
    reason = ""
    if not calling_enabled:
        reason = "Outbound calling is disabled (OUTBOUND_CALLING_ENABLED=false)"
    elif not allowed:
        try:
            from zoneinfo import ZoneInfo
            now_tz = datetime.now(tz=ZoneInfo(tz_name))
        except Exception:
            now_tz = datetime.now()
        day_abbr = now_tz.strftime("%a").lower()
        if day_abbr not in allowed_days:
            reason = f"Today ({now_tz.strftime('%A')}) is not an allowed calling day"
        else:
            reason = f"Current time is outside allowed window ({start_str}–{end_str} {tz_name})"
    win_err = await _outbound_window_error() if not allowed else None
    return {
        "allowed": allowed,
        "calling_enabled": calling_enabled,
        "reason": reason if not allowed else "Within allowed outbound calling window",
        "timezone": tz_name,
        "start_time": start_str,
        "end_time": end_str,
        "allowed_days": sorted(allowed_days),
        "next_allowed_at": win_err.get("next_allowed_at") if win_err else None,
    }
